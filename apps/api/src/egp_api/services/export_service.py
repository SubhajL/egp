"""Excel export service for backwards-compatible project list export."""

from __future__ import annotations

from io import BytesIO
from decimal import Decimal
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from egp_api.services.entitlement_service import TenantEntitlementService
from egp_shared_types.enums import ArtifactBucket, ClosedReason, NotificationType, ProjectState

if TYPE_CHECKING:
    from egp_db.repositories.document_repo import SqlDocumentRepository
    from egp_db.repositories.project_repo import ProjectDetail, ProjectRecord, SqlProjectRepository
    from egp_notifications.dispatcher import NotificationDispatcher


HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

COLUMNS = [
    ("download_date", 18),
    ("project_name", 45),
    ("organization", 30),
    ("project_number", 20),
    ("budget", 18),
    ("proposal_submission_date", 24),
    ("keyword", 20),
    ("tor_downloaded", 16),
    ("prelim_pricing", 16),
    ("search_name", 35),
    ("tracking_status", 24),
    ("closed_reason", 20),
    ("artifact_bucket", 24),
]


def _last_alias_value(detail: ProjectDetail | None, alias_type: str) -> str | None:
    if detail is None:
        return None
    for alias in reversed(detail.aliases):
        if alias.alias_type == alias_type and alias.alias_value.strip():
            return alias.alias_value
    return None


def _latest_keyword(detail: ProjectDetail | None) -> str | None:
    if detail is None:
        return None
    for event in reversed(detail.status_events):
        raw_snapshot = event.raw_snapshot
        if not isinstance(raw_snapshot, dict):
            continue
        keyword = str(raw_snapshot.get("keyword") or "").strip()
        if keyword:
            return keyword
    return None


def _derive_artifact_bucket(
    *,
    document_repository: SqlDocumentRepository | None,
    tenant_id: str,
    project_id: str,
) -> ArtifactBucket:
    if document_repository is None:
        return ArtifactBucket.NO_ARTIFACT_EVIDENCE
    return document_repository.get_artifact_bucket(tenant_id, project_id)


def _derive_tor_downloaded(
    *,
    project: ProjectRecord,
    artifact_bucket: ArtifactBucket,
) -> str:
    if artifact_bucket is ArtifactBucket.FINAL_TOR_DOWNLOADED:
        return "Yes"
    if project.project_state is ProjectState.TOR_DOWNLOADED:
        return "Yes"
    return "No"


def _derive_prelim_pricing(project: ProjectRecord) -> str:
    if project.project_state is ProjectState.PRELIM_PRICING_SEEN:
        return "Yes"
    if project.closed_reason is ClosedReason.PRELIM_PRICING:
        return "Yes"
    return "No"


def _project_to_row(
    *,
    project: ProjectRecord,
    detail: ProjectDetail | None,
    artifact_bucket: ArtifactBucket,
) -> list[str | float | None]:
    budget = float(project.budget_amount) if project.budget_amount else None
    return [
        project.last_seen_at[:10],
        project.project_name,
        project.organization_name,
        project.project_number,
        budget,
        project.proposal_submission_date,
        _latest_keyword(detail),
        _derive_tor_downloaded(project=project, artifact_bucket=artifact_bucket),
        _derive_prelim_pricing(project),
        _last_alias_value(detail, "search_name"),
        project.project_state.value,
        project.closed_reason.value if project.closed_reason else None,
        artifact_bucket.value,
    ]


class ExportService:
    def __init__(
        self,
        project_repository: SqlProjectRepository,
        *,
        document_repository: SqlDocumentRepository | None = None,
        entitlement_service: TenantEntitlementService | None = None,
        notification_dispatcher: NotificationDispatcher | None = None,
    ) -> None:
        self._repository = project_repository
        self._document_repository = document_repository
        self._entitlement_service = entitlement_service
        self._notification_dispatcher = notification_dispatcher

    def export_to_excel(
        self,
        *,
        tenant_id: str,
        project_states: list[str] | None = None,
        procurement_types: list[str] | None = None,
        closed_reasons: list[str] | None = None,
        organization: str | None = None,
        keyword: str | None = None,
        budget_min: Decimal | float | int | str | None = None,
        budget_max: Decimal | float | int | str | None = None,
        updated_after: str | None = None,
        has_changed_tor: bool | None = None,
        has_winner: bool | None = None,
    ) -> bytes:
        if self._entitlement_service is not None:
            self._entitlement_service.require_capability(
                tenant_id=tenant_id,
                capability="exports",
            )
        page = self._repository.list_projects(
            tenant_id=tenant_id,
            project_states=project_states,
            procurement_types=procurement_types,
            closed_reasons=closed_reasons,
            organization=organization,
            keyword=keyword,
            budget_min=budget_min,
            budget_max=budget_max,
            updated_after=updated_after,
            has_changed_tor=has_changed_tor,
            has_winner=has_winner,
            limit=10000,
            offset=0,
        )
        projects = page.items

        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = "โครงการ"

        # Header row
        for col_idx, (header, width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
            ws.column_dimensions[cell.column_letter].width = width

        # Data rows
        for row_idx, project in enumerate(projects, start=2):
            detail = self._repository.get_project_detail(
                tenant_id=tenant_id,
                project_id=project.id,
            )
            artifact_bucket = _derive_artifact_bucket(
                document_repository=self._document_repository,
                tenant_id=tenant_id,
                project_id=project.id,
            )
            row_data = _project_to_row(
                project=project,
                detail=detail,
                artifact_bucket=artifact_bucket,
            )
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if col_idx == 5 and isinstance(value, (int, float)):
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.freeze_panes = "A2"

        buffer = BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        if self._notification_dispatcher is not None:
            self._notification_dispatcher.dispatch(
                tenant_id=tenant_id,
                notification_type=NotificationType.EXPORT_READY,
            )
        return excel_bytes
