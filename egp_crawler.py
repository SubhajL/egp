#!/usr/bin/env python3
"""
e-GP Thailand Government Procurement Crawler

Searches gprocurement.go.th for government projects matching specified keywords,
downloads TOR documents, and logs projects to an Excel spreadsheet.

Launches your real Chrome browser and connects Playwright to it via CDP
(Chrome DevTools Protocol). Since it's your actual Chrome, Cloudflare
Turnstile treats it as a normal browser and passes automatically.

Usage:
    source .venv/bin/activate
    python egp_crawler.py                  # TOR profile (default)
    python egp_crawler.py --profile tor    # TOR: IT/data projects → OneDrive/Download/TOR
    python egp_crawler.py --profile toe    # TOE: AV/display equipment → OneDrive/Download/TOE
    python egp_crawler.py --profile lue    # LUE: PR/ads/training/web/online → OneDrive/Download/LUE

Configuration (optional environment variables):
    EGP_ENV_FILE                 Path to .env file (default: ./ .env or script dir)
    EGP_CHROME_PATH              Chrome executable path
    EGP_CDP_PORT                 Chrome remote debugging port (default 9222)
    EGP_DOWNLOAD_DIR             Where project folders/files are stored
    EGP_EXCEL_PATH               Full path to project_list.xlsx
    EGP_BROWSER_PROFILE_DIR      Dedicated Chrome profile dir (keep out of OneDrive)
    EGP_LOCAL_FALLBACK_DIR       Local fallback dir if OneDrive is inaccessible
    EGP_TEMP_DOWNLOAD_DIR        Temp download dir (defaults to system temp)
    EGP_MAIN_PAGE_URL            e-GP landing page URL
    EGP_SEARCH_URL               e-GP search URL
    EGP_KEYWORDS                 Comma/newline-separated keywords list
    EGP_MAX_PAGES_PER_KEYWORD    Pagination cap per keyword (default 15)
    EGP_NAV_TIMEOUT_MS           Page navigation timeout in ms (default 60000)
    EGP_CLOUDFLARE_TIMEOUT_MS    Cloudflare wait timeout in ms (default 120000)
    EGP_CLOUDFLARE_RELOAD_RETRIES  Page reload retries after Cloudflare timeout
    EGP_SEARCH_PAGE_RECOVERY_RETRIES  Fresh search-page retries after CF failure
    EGP_DOWNLOAD_TIMEOUT_MS      Ctrl+S/new-tab save timeout in ms (default 30000)
    EGP_SUBPAGE_DOWNLOAD_TIMEOUT_MS  Large TOR file timeout in ms (default 90000)
    EGP_DOWNLOAD_EVENT_TIMEOUT_MS    Wait for download event start in ms (default 15000)
"""

import argparse
import base64
import json
import builtins
import inspect
import os
import re
import shutil
import signal
import socket
import subprocess
import tempfile
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from threading import Lock
from urllib.parse import parse_qs, unquote, urljoin, urlparse

from egp_document_classifier import derive_artifact_bucket
from egp_shared_types.enums import ArtifactBucket, ClosedReason, ProcurementType, ProjectState
from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------


def env_get_str(var: str, default: str) -> str:
    value = os.getenv(var)
    if value is None:
        return default
    value = value.strip()
    return value if value else default


def env_get_int(
    var: str,
    default: int,
    *,
    min_value: int | None = None,
    max_value: int | None = None,
) -> int:
    raw = os.getenv(var)
    if raw is None:
        return default
    raw = raw.strip()
    if not raw:
        return default
    try:
        n = int(raw)
    except ValueError:
        return default
    if min_value is not None and n < min_value:
        return default
    if max_value is not None and n > max_value:
        return default
    return n


def env_get_float(
    var: str,
    default: float,
    *,
    min_value: float | None = None,
    max_value: float | None = None,
) -> float:
    raw = os.getenv(var)
    if raw is None:
        return default
    raw = raw.strip()
    if not raw:
        return default
    try:
        n = float(raw)
    except ValueError:
        return default
    if min_value is not None and n < min_value:
        return default
    if max_value is not None and n > max_value:
        return default
    return n


def env_get_path(var: str, default: Path) -> Path:
    raw = os.getenv(var)
    if raw is None:
        return default
    raw = raw.strip()
    if not raw:
        return default
    return Path(raw).expanduser()


def parse_keywords(text: str) -> list[str]:
    """Parse comma/newline-separated keywords string into a clean list."""
    if not text:
        return []
    parts = re.split(r"[,\n\r]+", text)
    cleaned = [p.strip() for p in parts if p and p.strip()]
    return cleaned


def keywords_from_env(var: str, default_keywords: list[str]) -> list[str]:
    raw = os.getenv(var)
    if raw is None:
        return list(default_keywords)
    parsed = parse_keywords(raw)
    return parsed if parsed else list(default_keywords)


def _strip_dotenv_inline_comment(value: str) -> str:
    """Strip inline comments for unquoted .env values (best-effort)."""
    v = value.strip()
    if not v or v[0] in ("'", '"'):
        return value
    # Treat " # ..." as comment, but keep literal hashes in values like URLs.
    if " #" in v:
        return v.split(" #", 1)[0].rstrip()
    return v


def _unquote_dotenv_value(value: str) -> str:
    v = value.strip()
    if len(v) >= 2 and v[0] == v[-1] and v[0] in ("'", '"'):
        inner = v[1:-1]
        # Minimal escape support
        inner = inner.replace("\\n", "\n").replace("\\r", "\r").replace("\\t", "\t")
        inner = inner.replace('\\"', '"').replace("\\'", "'").replace("\\\\", "\\")
        return inner
    return _strip_dotenv_inline_comment(v)


def load_dotenv_file(path: Path, *, override: bool = False) -> dict[str, str]:
    """Load KEY=VALUE pairs from a .env file into os.environ.

    This intentionally supports a conservative subset of dotenv syntax.
    Returns the dict of parsed key/value pairs.
    """
    parsed: dict[str, str] = {}
    try:
        content = path.read_text(encoding="utf-8")
    except Exception:
        return parsed

    for raw_line in content.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        m = re.match(
            r"^\s*(?:export\s+)?([A-Za-z_][A-Za-z0-9_]*)\s*=\s*(.*)\s*$", raw_line
        )
        if not m:
            continue
        key = m.group(1)
        value = _unquote_dotenv_value(m.group(2))
        parsed[key] = value
        if override:
            os.environ[key] = value
        else:
            os.environ.setdefault(key, value)

    return parsed


def load_dotenv_from_default_locations(*, override: bool = False) -> Path | None:
    """Load .env from EGP_ENV_FILE or default locations, returning the loaded path."""
    raw = os.getenv("EGP_ENV_FILE", "").strip()
    candidates: list[Path] = []
    if raw:
        candidates.append(Path(raw).expanduser())
    else:
        # Prefer script directory, then current working directory.
        candidates.append(Path(__file__).with_name(".env"))
        candidates.append(Path.cwd() / ".env")

    for path in candidates:
        try:
            if path.exists() and path.is_file():
                load_dotenv_file(path, override=override)
                return path
        except Exception:
            continue
    return None


def apply_env_config_overrides() -> None:
    """Re-apply all EGP_* environment variables to runtime globals.

    This is used after loading a .env file so we can override the module-level
    defaults without requiring the env vars to be present at import time.
    """
    global KEYWORDS
    global MAIN_PAGE_URL, SEARCH_URL
    global \
        DOWNLOAD_DIR, \
        EXCEL_PATH, \
        BROWSER_PROFILE_DIR, \
        TEMP_DOWNLOAD_DIR, \
        LOCAL_FALLBACK_DIR
    global \
        NAV_TIMEOUT, \
        CLOUDFLARE_TIMEOUT, \
        CLOUDFLARE_RELOAD_RETRIES, \
        SEARCH_PAGE_RECOVERY_RETRIES, \
        DOWNLOAD_TIMEOUT, \
        SUBPAGE_DOWNLOAD_TIMEOUT, \
        DOWNLOAD_EVENT_TIMEOUT
    global \
        DOWNLOAD_CLICK_RETRIES, \
        EXCEL_ACCESS_RETRIES, \
        EXCEL_ACCESS_RETRY_SECONDS, \
        MAX_PAGES_PER_KEYWORD
    global CHROME_PATH, CDP_PORT

    KEYWORDS = keywords_from_env("EGP_KEYWORDS", KEYWORDS_DEFAULT)
    MAIN_PAGE_URL = env_get_str("EGP_MAIN_PAGE_URL", MAIN_PAGE_URL)
    SEARCH_URL = env_get_str("EGP_SEARCH_URL", SEARCH_URL)

    DOWNLOAD_DIR = env_get_path("EGP_DOWNLOAD_DIR", DOWNLOAD_DIR)
    BROWSER_PROFILE_DIR = env_get_path("EGP_BROWSER_PROFILE_DIR", BROWSER_PROFILE_DIR)
    TEMP_DOWNLOAD_DIR = env_get_path("EGP_TEMP_DOWNLOAD_DIR", TEMP_DOWNLOAD_DIR)
    LOCAL_FALLBACK_DIR = env_get_path("EGP_LOCAL_FALLBACK_DIR", LOCAL_FALLBACK_DIR)

    EXCEL_PATH = env_get_path("EGP_EXCEL_PATH", DOWNLOAD_DIR / "project_list.xlsx")

    NAV_TIMEOUT = env_get_int("EGP_NAV_TIMEOUT_MS", NAV_TIMEOUT, min_value=5_000)
    CLOUDFLARE_TIMEOUT = env_get_int(
        "EGP_CLOUDFLARE_TIMEOUT_MS",
        CLOUDFLARE_TIMEOUT,
        min_value=10_000,
    )
    CLOUDFLARE_RELOAD_RETRIES = env_get_int(
        "EGP_CLOUDFLARE_RELOAD_RETRIES",
        CLOUDFLARE_RELOAD_RETRIES,
        min_value=0,
        max_value=5,
    )
    SEARCH_PAGE_RECOVERY_RETRIES = env_get_int(
        "EGP_SEARCH_PAGE_RECOVERY_RETRIES",
        SEARCH_PAGE_RECOVERY_RETRIES,
        min_value=0,
        max_value=5,
    )
    DOWNLOAD_TIMEOUT = env_get_int(
        "EGP_DOWNLOAD_TIMEOUT_MS", DOWNLOAD_TIMEOUT, min_value=5_000
    )
    SUBPAGE_DOWNLOAD_TIMEOUT = env_get_int(
        "EGP_SUBPAGE_DOWNLOAD_TIMEOUT_MS",
        SUBPAGE_DOWNLOAD_TIMEOUT,
        min_value=10_000,
    )
    DOWNLOAD_EVENT_TIMEOUT = env_get_int(
        "EGP_DOWNLOAD_EVENT_TIMEOUT_MS",
        DOWNLOAD_EVENT_TIMEOUT,
        min_value=2_000,
    )
    DOWNLOAD_CLICK_RETRIES = env_get_int(
        "EGP_DOWNLOAD_CLICK_RETRIES",
        DOWNLOAD_CLICK_RETRIES,
        min_value=0,
        max_value=10,
    )
    EXCEL_ACCESS_RETRIES = env_get_int(
        "EGP_EXCEL_ACCESS_RETRIES",
        EXCEL_ACCESS_RETRIES,
        min_value=0,
        max_value=20,
    )
    EXCEL_ACCESS_RETRY_SECONDS = env_get_float(
        "EGP_EXCEL_ACCESS_RETRY_SECONDS",
        EXCEL_ACCESS_RETRY_SECONDS,
        min_value=0.0,
        max_value=60.0,
    )
    MAX_PAGES_PER_KEYWORD = env_get_int(
        "EGP_MAX_PAGES_PER_KEYWORD",
        MAX_PAGES_PER_KEYWORD,
        min_value=1,
        max_value=200,
    )

    CHROME_PATH = env_get_str("EGP_CHROME_PATH", CHROME_PATH)
    CDP_PORT = env_get_int("EGP_CDP_PORT", CDP_PORT, min_value=1, max_value=65535)


KEYWORDS_DEFAULT = [
    "วิเคราะห์ข้อมูล",
    "ระบบสารสนเทศ",
    "เทคโนโลยีสารสนเทศ",
    "ระบบคลังข้อมูล",
    "ระบบฐานข้อมูลใหญ่",
    "แผนแม่บท",
    "แผนปฏิบัติการ",
    "สถาปัตยกรรมองค์กร",
    "ธรรมาภิบาลข้อมูล",
    "ที่ปรึกษา",
    "ระบบวิเคราะห์",
    "ระบบบริหารจัดการ",
]

KEYWORDS_TOE_DEFAULT = [
    "จอแสดงผล",
    "จอแสดงภาพ",
    "interactive",
    "LED Wall",
    "Smart TV",
]

KEYWORDS_LUE_DEFAULT = [
    "ประชาสัมพันธ์",
    "โฆษณา",
    "จัดอบรม",
    "เว็บไซต์",
    "สื่อออนไลน์",
]

# Profile registry: maps profile name → keyword list + storage paths.
# Add new profiles here; no other code changes required.
PROFILE_DEFAULTS: dict[str, dict] = {
    "tor": {
        "keywords": KEYWORDS_DEFAULT,
        "download_dir": Path.home() / "OneDrive" / "Download" / "TOR",
        "local_fallback_dir": Path.home() / "download" / "TOR",
    },
    "toe": {
        "keywords": KEYWORDS_TOE_DEFAULT,
        "download_dir": Path.home() / "OneDrive" / "Download" / "TOE",
        "local_fallback_dir": Path.home() / "download" / "TOE",
    },
    "lue": {
        "keywords": KEYWORDS_LUE_DEFAULT,
        "download_dir": Path.home() / "OneDrive" / "Download" / "LUE",
        "local_fallback_dir": Path.home() / "download" / "LUE",
    },
}


def apply_profile_defaults(profile: str) -> None:
    """Seed EGP_* env vars from the selected profile using setdefault.

    Called at the very start of main(), before .env loading and
    apply_env_config_overrides(), so the full priority chain is preserved:
      shell env var  >  --profile flag  >  .env file  >  hardcoded default
    """
    cfg = PROFILE_DEFAULTS[profile]
    os.environ.setdefault("EGP_KEYWORDS", ",".join(cfg["keywords"]))
    os.environ.setdefault("EGP_DOWNLOAD_DIR", str(cfg["download_dir"]))
    os.environ.setdefault("EGP_LOCAL_FALLBACK_DIR", str(cfg["local_fallback_dir"]))


KEYWORDS = keywords_from_env("EGP_KEYWORDS", KEYWORDS_DEFAULT)

MAIN_PAGE_URL = env_get_str(
    "EGP_MAIN_PAGE_URL",
    "https://www.gprocurement.go.th/new_index.html",
)
SEARCH_URL = env_get_str(
    "EGP_SEARCH_URL",
    "https://process5.gprocurement.go.th/egp-agpc01-web/announcement",
)
DOWNLOAD_DIR = env_get_path(
    "EGP_DOWNLOAD_DIR", Path.home() / "OneDrive" / "Download" / "TOR"
)
# Browser profile lives OUTSIDE OneDrive to avoid sync popups
BROWSER_PROFILE_DIR = env_get_path(
    "EGP_BROWSER_PROFILE_DIR",
    Path.home() / "download" / "TOR" / ".browser_profile",
)
# Temp download dir outside OneDrive — files are moved to DOWNLOAD_DIR after completion
TEMP_DOWNLOAD_DIR = env_get_path(
    "EGP_TEMP_DOWNLOAD_DIR",
    Path(tempfile.gettempdir()) / "egp_downloads",
)
EXCEL_PATH = env_get_path("EGP_EXCEL_PATH", DOWNLOAD_DIR / "project_list.xlsx")

TARGET_STATUS = "หนังสือเชิญชวน/ประกาศเชิญชวน"
SKIP_TEXT = "สรุปข้อมูลการเสนอราคาเบื้องต้น"
SKIP_KEYWORDS_IN_PROJECT = [
    "ทางหลวง",
    "วิธีคัดเลือก",
    "บำรุงรักษา",
]  # Skip projects containing these in name or org
NEXT_PAGE_SELECTOR = (
    "a:has-text('ถัดไป'), "
    "button:has-text('ถัดไป'), "
    "button[aria-label='next'], "
    "a:has-text('»'), "
    "li.next:not(.disabled) a"
)

# Documents to download from project info page
DOCS_TO_DOWNLOAD = [
    "ประกาศเชิญชวน",
    "ประกาศราคากลาง",
    "ร่างเอกสารประกวดราคา",
    "เอกสารประกวดราคา",
]

# Alternate labels seen for TOR-like docs (especially consulting procurements)
TOR_DOC_MATCH_TERMS = [
    "ร่างเอกสารประกวดราคา",
    "เอกสารประกวดราคา",
    "ร่างขอบเขตของงาน",
    "ขอบเขตของงาน",
    "เอกสารจ้างที่ปรึกษา",
    "terms of reference",
    "tor",
]
DRAFT_TOR_DOC_MATCH_TERMS = [
    "ร่างขอบเขตของงาน",
    "ร่างเอกสารประกวดราคา",
    "draft tor",
    "draft terms of reference",
]

EXCEL_HEADERS = [
    "download_date",
    "project_name",
    "organization",
    "project_number",
    "budget",
    "proposal_submission_date",
    "keyword",
    "tor_downloaded",
    "prelim_pricing",
    "search_name",
    "tracking_status",
    "closed_reason",
    "artifact_bucket",
]


@dataclass(frozen=True, slots=True)
class ProjectDocumentSummary:
    saved_labels: tuple[str, ...]
    artifact_bucket: ArtifactBucket

    @property
    def tor_downloaded(self) -> bool:
        return self.artifact_bucket is ArtifactBucket.FINAL_TOR_DOWNLOADED

    @property
    def draft_tor_downloaded(self) -> bool:
        return self.artifact_bucket is ArtifactBucket.DRAFT_PLUS_PRICING


@dataclass(frozen=True, slots=True)
class KeywordResumeState:
    keyword_index: int
    keyword: str
    page_num: int = 1


# Timeouts (ms)
NAV_TIMEOUT = env_get_int("EGP_NAV_TIMEOUT_MS", 60_000, min_value=5_000)
CLOUDFLARE_TIMEOUT = env_get_int("EGP_CLOUDFLARE_TIMEOUT_MS", 120_000, min_value=10_000)
CLOUDFLARE_RELOAD_RETRIES = env_get_int(
    "EGP_CLOUDFLARE_RELOAD_RETRIES", 1, min_value=0, max_value=5
)
SEARCH_PAGE_RECOVERY_RETRIES = env_get_int(
    "EGP_SEARCH_PAGE_RECOVERY_RETRIES", 1, min_value=0, max_value=5
)
DOWNLOAD_TIMEOUT = env_get_int("EGP_DOWNLOAD_TIMEOUT_MS", 30_000, min_value=5_000)
SUBPAGE_DOWNLOAD_TIMEOUT = env_get_int(
    "EGP_SUBPAGE_DOWNLOAD_TIMEOUT_MS", 90_000, min_value=10_000
)
DOWNLOAD_EVENT_TIMEOUT = env_get_int(
    "EGP_DOWNLOAD_EVENT_TIMEOUT_MS", 15_000, min_value=2_000
)
DOWNLOAD_CLICK_RETRIES = env_get_int(
    "EGP_DOWNLOAD_CLICK_RETRIES", 2, min_value=0, max_value=10
)
EXCEL_ACCESS_RETRIES = env_get_int(
    "EGP_EXCEL_ACCESS_RETRIES", 3, min_value=0, max_value=20
)
EXCEL_ACCESS_RETRY_SECONDS = env_get_float(
    "EGP_EXCEL_ACCESS_RETRY_SECONDS", 1.0, min_value=0.0, max_value=60.0
)

MAX_PAGES_PER_KEYWORD = env_get_int(
    "EGP_MAX_PAGES_PER_KEYWORD", 15, min_value=1, max_value=200
)
RUNTIME_LOG_FILENAME = "egp_crawler_runtime.log"
LOCAL_FALLBACK_DIR = env_get_path(
    "EGP_LOCAL_FALLBACK_DIR", Path.home() / "download" / "TOR"
)
ALLOWED_DOWNLOAD_HOST_SUFFIXES = ("gprocurement.go.th",)
_REAL_PRINT = builtins.print
_REAL_SLEEP = time.sleep
_LOG_LOCK = Lock()
_USING_LOCAL_EXCEL_FALLBACK = False
_ONEDRIVE_EXCEL_PATH: Path | None = None
_TARGET_STATUS_COMPACT = re.sub(r"\s+", "", TARGET_STATUS)


def _runtime_log_paths() -> list[Path]:
    """Preferred + fallback log paths."""
    return [
        Path(EXCEL_PATH).with_name(RUNTIME_LOG_FILENAME),
        Path.cwd() / RUNTIME_LOG_FILENAME,
    ]


def _append_runtime_log(line: str) -> None:
    for path in _runtime_log_paths():
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            with _LOG_LOCK:
                with path.open("a", encoding="utf-8") as fh:
                    fh.write(line + "\n")
            return
        except Exception:
            continue


def _infer_log_level(message: str) -> str:
    lowered = (message or "").lower()
    if any(k in lowered for k in (" error", "error:", "exception", "traceback")):
        return "ERROR"
    if any(k in lowered for k in ("warning", "timeout", "failed", "could not")):
        return "WARN"
    return "INFO"


def _log_message(message: str, level: str | None = None) -> None:
    lvl = level or _infer_log_level(message)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _append_runtime_log(f"{ts} [{lvl}] {message}")


def print(*args, **kwargs):  # type: ignore[override]
    """Mirror all console prints to runtime log file."""
    _REAL_PRINT(*args, **kwargs)
    try:
        sep = kwargs.get("sep", " ")
        msg = sep.join(str(a) for a in args)
        _log_message(msg)
    except Exception:
        pass


def logged_sleep(seconds: float, reason: str = "") -> None:
    """Sleep while logging pause location and duration."""
    try:
        frame = inspect.stack()[1]
        location = f"{Path(frame.filename).name}:{frame.lineno}:{frame.function}"
    except Exception:
        location = "unknown"
    msg = f"Pause {seconds:.1f}s at {location}"
    if reason:
        msg += f" ({reason})"
    _log_message(msg, level="PAUSE")
    _REAL_SLEEP(seconds)


def _can_read_file_with_retry(path: Path, retries: int = EXCEL_ACCESS_RETRIES) -> bool:
    """Return True if file is readable in this process (or doesn't exist yet)."""
    if not path.exists():
        return True
    for attempt in range(retries + 1):
        try:
            with path.open("rb") as fh:
                fh.read(8)
            return True
        except PermissionError:
            if attempt < retries:
                _log_message(
                    f"Excel read permission blocked for {path}; retrying ({attempt + 1}/{retries})",
                    level="WARN",
                )
                logged_sleep(EXCEL_ACCESS_RETRY_SECONDS, "excel read retry")
                continue
            return False
        except Exception:
            return False
    return False


def _can_write_dir(path: Path) -> bool:
    """Return True if path is writable in this process."""
    try:
        path.mkdir(parents=True, exist_ok=True)
        probe = path / ".egp_write_probe.tmp"
        probe.write_text("ok", encoding="utf-8")
        probe.unlink(missing_ok=True)
        return True
    except Exception:
        return False


def _safe_copy_file(src: Path, dst: Path, label: str) -> bool:
    """Best-effort file copy with logging, returns True on success."""
    try:
        if not src.exists():
            return False
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dst)
        _log_message(f"{label}: copied {src} -> {dst}", level="INFO")
        return True
    except Exception as e:
        _log_message(f"{label}: copy failed {src} -> {dst}: {e}", level="WARN")
        return False


def _mirror_to_local_fallback(path: Path) -> Path:
    """Map a OneDrive path to local fallback while preserving relative structure."""
    try:
        rel = path.relative_to(DOWNLOAD_DIR)
        return LOCAL_FALLBACK_DIR / rel
    except Exception:
        return LOCAL_FALLBACK_DIR / path.name


def _safe_remove_dir(path: Path, label: str) -> None:
    """Best-effort folder removal that logs and continues on failure."""
    try:
        if path.exists():
            shutil.rmtree(path)
            print(f"      Deleted folder: {path.name}")
    except Exception as e:
        _log_message(f"{label}: could not delete folder {path}: {e}", level="WARN")


def configure_runtime_paths() -> None:
    """Configure runtime storage paths with fallback when OneDrive is inaccessible."""
    global DOWNLOAD_DIR, EXCEL_PATH, _USING_LOCAL_EXCEL_FALLBACK, _ONEDRIVE_EXCEL_PATH

    _USING_LOCAL_EXCEL_FALLBACK = False
    one_drive_download = DOWNLOAD_DIR
    one_drive_excel = one_drive_download / EXCEL_PATH.name
    _ONEDRIVE_EXCEL_PATH = one_drive_excel

    if not _can_write_dir(one_drive_download):
        fallback_download = LOCAL_FALLBACK_DIR
        _log_message(
            f"WARNING: cannot write download directory {one_drive_download}; using {fallback_download}",
            level="WARN",
        )
        DOWNLOAD_DIR = fallback_download
        EXCEL_PATH = DOWNLOAD_DIR / EXCEL_PATH.name

    if not _can_read_file_with_retry(one_drive_excel):
        fallback_excel = LOCAL_FALLBACK_DIR / one_drive_excel.name
        _log_message(
            f"WARNING: cannot read Excel file {one_drive_excel}; using local fallback {fallback_excel}",
            level="WARN",
        )
        # If OneDrive becomes readable during retries, seed local copy from latest.
        if _can_read_file_with_retry(one_drive_excel, retries=1):
            _safe_copy_file(one_drive_excel, fallback_excel, "Excel fallback seed")
        EXCEL_PATH = fallback_excel
        _USING_LOCAL_EXCEL_FALLBACK = True

    # Final guard: ensure fallback parent is writable.
    if not _can_write_dir(EXCEL_PATH.parent):
        emergency_excel = Path.cwd() / EXCEL_PATH.name
        _log_message(
            f"WARNING: cannot write Excel parent {EXCEL_PATH.parent}; using {emergency_excel}",
            level="WARN",
        )
        EXCEL_PATH = emergency_excel


def sync_excel_back_to_onedrive() -> None:
    """If fallback Excel was used, try syncing it back to OneDrive."""
    if not _USING_LOCAL_EXCEL_FALLBACK:
        return
    if not _ONEDRIVE_EXCEL_PATH:
        return
    if EXCEL_PATH == _ONEDRIVE_EXCEL_PATH:
        return
    if not EXCEL_PATH.exists():
        _log_message(
            "Excel sync skipped: local fallback file does not exist", level="WARN"
        )
        return
    if not _can_write_dir(_ONEDRIVE_EXCEL_PATH.parent):
        _log_message(
            f"Excel sync skipped: OneDrive directory not writable {_ONEDRIVE_EXCEL_PATH.parent}",
            level="WARN",
        )
        return
    _safe_copy_file(EXCEL_PATH, _ONEDRIVE_EXCEL_PATH, "Excel fallback sync")


# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------


STALE_ANNOUNCEMENT_DAYS = 45  # Delete folders for announcements older than this
SITE_ERROR_TOAST_TEXT = "ระบบเกิดข้อผิดพลาด"
SITE_ERROR_TOAST_HINT = "กรุณาตรวจสอบ"
TOAST_RECOVERY_RETRIES = 2


def parse_buddhist_date(text: str) -> "date | None":
    """Parse a Thai Buddhist-era date string (DD/MM/YYYY+543) to a Gregorian date.

    Returns None if the string cannot be parsed.
    """
    text = text.strip()
    match = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", text)
    if not match:
        return None
    day, month, buddhist_year = (
        int(match.group(1)),
        int(match.group(2)),
        int(match.group(3)),
    )
    try:
        return date(buddhist_year - 543, month, day)
    except ValueError:
        return None


def has_site_error_toast_text(text: str) -> bool:
    """Check if text looks like the runtime red error toast from e-GP site."""
    compact = re.sub(r"\s+", "", text or "")
    return SITE_ERROR_TOAST_TEXT in compact and SITE_ERROR_TOAST_HINT in compact


def is_tor_file(filename: str) -> bool:
    """Check if a downloaded filename is an actual TOR document (not a pricebuild).

    Files with 'pricebuild' prefix or 'pB' prefix (e.g. pB0.pdf) are price
    build-up documents, not TOR. Actual TOR files are the bidding specification
    documents (ขอบเขตของงาน / เอกสารประกวดราคา).
    """
    if not filename:
        return False
    # pricebuild_XXXX_YYYY.zip — price estimate summary
    lowered = filename.lower()
    if lowered.startswith("pricebuild"):
        return False
    # pB0.pdf, pB1.pdf, etc. — individual price build-up pages
    if re.match(r"^pb\d+\.pdf$", lowered):
        return False
    return True


def is_tor_doc_label(label: str) -> bool:
    """Check if a document row label likely refers to TOR-like documents."""
    lowered = label.strip().lower()
    return any(term in lowered for term in TOR_DOC_MATCH_TERMS)


def is_draft_tor_doc_label(label: str) -> bool:
    """Check if a document row label refers to a draft/public-hearing TOR."""
    lowered = label.strip().lower()
    return any(term in lowered for term in DRAFT_TOR_DOC_MATCH_TERMS)


def is_final_tor_doc_label(label: str) -> bool:
    """Check if a document row label refers to a final invitation-stage TOR."""
    return is_tor_doc_label(label) and not is_draft_tor_doc_label(label)


def extract_file_label_from_cell_texts(texts: list[str]) -> str:
    """Best-effort extraction of a file label from a table row's cell texts.

    Many e-GP download tables use a numeric first column ("ลำดับ") and put the
    actual filename in the next column, while other pages put the filename in
    the first column.
    """
    cleaned = [str(t or "").strip() for t in (texts or [])]
    cleaned = [t for t in cleaned if t]
    if not cleaned:
        return ""

    for t in cleaned:
        if re.search(r"\.(?:zip|pdf|docx?|xlsx?)\b", t, flags=re.IGNORECASE):
            return t

    if len(cleaned) >= 2 and cleaned[0].isdigit():
        return cleaned[1]
    return cleaned[0]


def pagination_button_is_disabled(
    aria_disabled: str | None, disabled: bool | None, class_name: str | None
) -> bool:
    if disabled is True:
        return True
    if aria_disabled and aria_disabled.strip().lower() == "true":
        return True
    if class_name and re.search(r"(?:^|\s)disabled(?:\s|$)", class_name):
        return True
    return False


def ensure_excel_headers(ws) -> None:
    """Backfill/repair header row so legacy sheets upgrade safely."""
    for idx, header in enumerate(EXCEL_HEADERS, start=1):
        if ws.cell(1, idx).value != header:
            ws.cell(1, idx).value = header


_TERMINAL_TRACKING_STATES = {
    ProjectState.TOR_DOWNLOADED.value,
    ProjectState.PRELIM_PRICING_SEEN.value,
    ProjectState.WINNER_ANNOUNCED.value,
    ProjectState.CONTRACT_SIGNED.value,
    ProjectState.CLOSED_TIMEOUT_CONSULTING.value,
    ProjectState.CLOSED_STALE_NO_TOR.value,
    ProjectState.CLOSED_MANUAL.value,
}


def is_terminal_tracking_status(value: str | None) -> bool:
    normalized = str(value or "").strip()
    if normalized in _TERMINAL_TRACKING_STATES:
        return True
    return False


def infer_procurement_type(project_name: str, organization: str = "") -> ProcurementType:
    combined = f"{project_name} {organization}"
    if "ที่ปรึกษา" in combined:
        return ProcurementType.CONSULTING
    return ProcurementType.SERVICES


def derive_tracking_status(
    *,
    project_name: str,
    organization: str,
    artifact_bucket: ArtifactBucket,
    prelim_pricing: bool = False,
    stale_without_tor: bool = False,
) -> tuple[ProjectState, ClosedReason | None]:
    if prelim_pricing:
        return (ProjectState.PRELIM_PRICING_SEEN, ClosedReason.PRELIM_PRICING)
    if stale_without_tor:
        return (ProjectState.CLOSED_STALE_NO_TOR, ClosedReason.STALE_NO_TOR)
    if artifact_bucket is ArtifactBucket.FINAL_TOR_DOWNLOADED:
        return (ProjectState.TOR_DOWNLOADED, None)
    if artifact_bucket is ArtifactBucket.DRAFT_PLUS_PRICING:
        return (ProjectState.OPEN_PUBLIC_HEARING, None)
    if infer_procurement_type(project_name, organization) is ProcurementType.CONSULTING:
        return (ProjectState.OPEN_CONSULTING, None)
    return (ProjectState.OPEN_INVITATION, None)


def write_project_manifest(
    *,
    project_dir: Path,
    project_info: dict[str, object],
    tracking_status: ProjectState,
    closed_reason: ClosedReason | None,
    artifact_bucket: ArtifactBucket,
) -> Path:
    project_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = project_dir / "crawler_manifest.json"
    payload = {
        "project_number": str(project_info.get("project_number") or ""),
        "project_name": str(project_info.get("project_name") or ""),
        "search_name": str(project_info.get("search_name") or ""),
        "keyword": str(project_info.get("keyword") or ""),
        "tracking_status": tracking_status.value,
        "closed_reason": closed_reason.value if closed_reason is not None else None,
        "artifact_bucket": artifact_bucket.value,
        "saved_files": sorted(
            path.name for path in project_dir.iterdir() if path.is_file() and path.name != manifest_path.name
        ),
        "written_at": datetime.now().isoformat(timespec="seconds"),
        "saved_by": "crawler",
    }
    manifest_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return manifest_path


def sanitize_dirname(name: str) -> str:
    """Remove characters illegal in directory names and truncate to 100 chars."""
    cleaned = re.sub(r'[\\/*?:"<>|]', "", name)
    cleaned = cleaned.replace("\n", " ").replace("\r", " ").strip()
    # Collapse multiple spaces
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned[:100]


def sanitize_filename(name: str) -> str:
    """Remove characters illegal in filenames (no truncation)."""
    cleaned = re.sub(r'[\\/*?:"<>|]', "", name)
    cleaned = cleaned.replace("\n", " ").replace("\r", " ").strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned


def sanitize_filename_preserve_suffix(name: str, max_len: int = 100) -> str:
    """Sanitize and truncate while preserving the full suffix (e.g. .tar.gz)."""
    cleaned = sanitize_filename(name)
    if len(cleaned) <= max_len:
        return cleaned

    suffixes = Path(cleaned).suffixes
    ext = "".join(suffixes)
    if ext and cleaned.endswith(ext):
        stem = cleaned[: -len(ext)]
    else:
        ext = ""
        stem = cleaned

    if not ext:
        return cleaned[:max_len]

    max_stem_len = max_len - len(ext)
    if max_stem_len < 1:
        return ext[-max_len:]
    stem = stem.strip()[:max_stem_len].rstrip()
    if not stem:
        stem = "file"[:max_stem_len]
    return f"{stem}{ext}"


def build_safe_filename(stem: str, ext: str | None, max_len: int = 100) -> str:
    """Build a safe filename from a stem and extension, truncating stem to fit."""
    safe_stem = sanitize_filename(stem)
    if not safe_stem:
        safe_stem = "file"

    safe_ext = (ext or "").strip()
    if safe_ext and not safe_ext.startswith("."):
        safe_ext = f".{safe_ext}"

    if not safe_ext:
        return safe_stem[:max_len]

    max_stem_len = max_len - len(safe_ext)
    if max_stem_len < 1:
        return safe_ext[-max_len:]
    return f"{safe_stem[:max_stem_len]}{safe_ext}"


def extract_document_url_from_viewer_url(url: str) -> str | None:
    """Extract the underlying document URL when Chrome shows a viewer page.

    Chrome's PDF viewer commonly uses a chrome-extension:// URL with the real
    file URL embedded as a `file=` query parameter.
    """
    if not url:
        return None
    try:
        parsed = urlparse(url)
        qs = parse_qs(parsed.query)
        candidates = (qs.get("file") or []) + (qs.get("url") or [])
        for raw in candidates:
            decoded = unquote(raw)
            if decoded.startswith(("http://", "https://")):
                return decoded

        for values in qs.values():
            for raw in values:
                decoded = unquote(raw)
                if decoded.startswith(("http://", "https://")):
                    return decoded
    except Exception:
        return None
    return None


def guess_extension_from_content_type(content_type: str | None) -> str | None:
    if not content_type:
        return None
    lowered = content_type.lower()
    if "application/pdf" in lowered:
        return ".pdf"
    if "application/zip" in lowered or "application/x-zip-compressed" in lowered:
        return ".zip"
    return None


def sniff_extension_from_bytes(data: bytes) -> str | None:
    if not data:
        return None
    stripped = data.lstrip()
    if stripped.startswith(b"%PDF-"):
        return ".pdf"
    if data.startswith((b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")):
        return ".zip"
    return None


def looks_like_html_bytes(data: bytes) -> bool:
    if not data:
        return False
    prefix = data.lstrip()[:2048].lower()
    if prefix.startswith(b"<!doctype html") or prefix.startswith(b"<html"):
        return True
    return b"<html" in prefix or b"<head" in prefix or b"<title" in prefix


def is_allowed_download_url(url: str) -> bool:
    """Return True if URL is safe/expected for request-based fallback downloads."""
    try:
        parsed = urlparse(url)
    except Exception:
        return False
    if parsed.scheme not in ("http", "https"):
        return False
    host = (parsed.hostname or "").strip().lower()
    if not host:
        return False
    if re.fullmatch(r"\d{1,3}(?:\.\d{1,3}){3}", host) or host == "localhost":
        return False
    return any(
        host == suf or host.endswith(f".{suf}")
        for suf in ALLOWED_DOWNLOAD_HOST_SUFFIXES
    )


def resolve_http_url(candidate: str | None, base_url: str | None = None) -> str | None:
    """Resolve relative URLs and return an absolute http(s) URL when possible."""
    if not candidate:
        return None
    raw = candidate.strip()
    if not raw:
        return None
    extracted = extract_document_url_from_viewer_url(raw)
    if extracted:
        raw = extracted
    if raw.startswith(("http://", "https://")):
        return raw
    if base_url and not raw.lower().startswith(
        ("javascript:", "data:", "blob:", "chrome:", "about:")
    ):
        try:
            resolved = urljoin(base_url, raw)
        except Exception:
            resolved = None
        if resolved and resolved.startswith(("http://", "https://")):
            return resolved
    return None


def is_show_htmlfile_url(url: str | None) -> bool:
    if not url:
        return False
    try:
        parsed = urlparse(url)
    except Exception:
        return "showhtmlfile" in url.lower()
    if "showhtmlfile" in (parsed.path or "").lower():
        return True
    qs = parse_qs(parsed.query)
    proc_id = (qs.get("proc_id") or qs.get("procId") or qs.get("procID") or [""])[0]
    return str(proc_id).lower() == "showhtmlfile"


def extract_url_from_onclick(
    onclick: str | None, base_url: str | None = None
) -> str | None:
    """Best-effort: extract a URL-like string from an onclick attribute."""
    if not onclick:
        return None
    text = str(onclick)

    # First, prefer absolute URLs.
    m = re.search(r"""https?://[^\s'")\\]+""", text, flags=re.IGNORECASE)
    if m:
        return resolve_http_url(m.group(0), base_url=base_url)

    # Then, look for a relative path beginning with /.
    m2 = re.search(r"""['"](/[^'"]+)['"]""", text)
    if m2:
        return resolve_http_url(m2.group(1), base_url=base_url)

    # Finally, look for a servlet/jsp-ish path without leading slash.
    m3 = re.search(
        r"""['"]([^'"]*(?:procsearch\.sch|ShowHTMLFile)[^'"]*)['"]""",
        text,
        flags=re.IGNORECASE,
    )
    if m3:
        return resolve_http_url(m3.group(1), base_url=base_url)

    return None


def save_show_htmlfile_as_file(
    page, project_dir: Path, doc_name: str, prefer_pdf: bool = True
) -> str | None:
    """Save a ShowHTMLFile HTML viewer as PDF (preferred) or HTML."""
    try:
        url = page.url
    except Exception:
        url = ""
    if not is_show_htmlfile_url(url):
        return None

    if prefer_pdf:
        try:
            cdp = page.context.new_cdp_session(page)
            try:
                cdp.send("Page.enable")
            except Exception:
                pass
            res = cdp.send(
                "Page.printToPDF",
                {
                    "printBackground": True,
                    "preferCSSPageSize": True,
                },
            )
            data_b64 = (res or {}).get("data")
            if data_b64:
                pdf_bytes = base64.b64decode(data_b64)
                return _save_bytes_to_project(
                    project_dir, build_safe_filename(doc_name, ".pdf"), pdf_bytes
                )
        except Exception as e:
            print(f"      ShowHTMLFile PDF save failed for {doc_name}: {e}")

    try:
        html = page.content()
    except Exception as e:
        print(f"      ShowHTMLFile HTML save failed for {doc_name}: {e}")
        return None
    return _save_bytes_to_project(
        project_dir, build_safe_filename(doc_name, ".html"), html.encode("utf-8")
    )


def filename_from_content_disposition(header_value: str | None) -> str | None:
    if not header_value:
        return None
    header = header_value.strip()

    # RFC 5987: filename*=UTF-8''...
    m = re.search(r"""filename\*\s*=\s*([^']*)''([^;]+)""", header, flags=re.IGNORECASE)
    if m:
        raw = m.group(2).strip().strip('"')
        try:
            return unquote(raw)
        except Exception:
            return raw

    m2 = re.search(r"""filename\s*=\s*\"?([^\";]+)\"?""", header, flags=re.IGNORECASE)
    if m2:
        return m2.group(1).strip()
    return None


def load_existing_projects(excel_path: Path | None = None) -> dict[str, bool]:
    """Load projects from Excel with their completion status.

    Returns dict mapping keys → fully_done (True/False).

    We prefer stable keys (`project_number`, `search_name`) and only fall back to
    `project_name` when neither is available. This avoids re-announcements with a
    reused title incorrectly inheriting a previous row's completion state.
    """
    excel_path = Path(excel_path) if excel_path else EXCEL_PATH
    if not excel_path.exists():
        return {}

    try:
        wb = load_workbook(excel_path)
    except PermissionError:
        _log_message(f"ERROR: cannot read Excel file {excel_path}", level="ERROR")
        return {}
    ws = wb.active
    existing: dict[str, bool] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        project_name = row[1]  # column B = project_name
        if not project_name:
            continue
        name = str(project_name).strip()
        # column H = tor_downloaded (index 7)
        tor_status = str(row[7]).strip().lower() if len(row) > 7 and row[7] else "no"
        # column I = prelim_pricing (index 8)
        prelim = str(row[8]).strip().lower() if len(row) > 8 and row[8] else "no"
        tracking_status = str(row[10]).strip() if len(row) > 10 and row[10] else ""
        done = (
            is_terminal_tracking_status(tracking_status)
            or tor_status == "yes"
            or prelim == "yes"
        )

        # Index by project_number (column D, index 3)
        proj_num = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        if proj_num:
            existing[proj_num] = done

        # Index by search_name (column J, index 9)
        search_name = str(row[9]).strip() if len(row) > 9 and row[9] else ""
        if search_name:
            existing[search_name] = done
        elif not proj_num:
            existing[name] = done

    return existing


def load_existing_project_row_stats(
    excel_path: Path | None = None,
) -> tuple[int, int, int]:
    """Return (total_rows, complete_rows, incomplete_rows) from Excel project rows.

    This is separate from the dedup key map, which intentionally stores multiple
    keys per project (name/number/search_name) and should not be used for row counts.
    """
    excel_path = Path(excel_path) if excel_path else EXCEL_PATH
    if not excel_path.exists():
        return (0, 0, 0)

    try:
        wb = load_workbook(excel_path)
    except PermissionError:
        _log_message(f"ERROR: cannot read Excel stats file {excel_path}", level="ERROR")
        return (0, 0, 0)
    ws = wb.active
    total_rows = 0
    complete_rows = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        project_name = row[1] if len(row) > 1 else None
        if not project_name:
            continue
        total_rows += 1
        tor_status = str(row[7]).strip().lower() if len(row) > 7 and row[7] else "no"
        prelim = str(row[8]).strip().lower() if len(row) > 8 and row[8] else "no"
        tracking_status = str(row[10]).strip() if len(row) > 10 and row[10] else ""
        if (
            is_terminal_tracking_status(tracking_status)
            or tor_status == "yes"
            or prelim == "yes"
        ):
            complete_rows += 1

    return (total_rows, complete_rows, total_rows - complete_rows)


def update_excel(project_info: dict, excel_path: Path | None = None) -> None:
    """Create or append to the project list Excel file.

    Re-announcements must not merge by shared title text alone. Prefer exact
    `project_number`, then exact `search_name`, and only fall back to name-only
    rows that never had a project number.
    """
    excel_path = Path(excel_path) if excel_path else EXCEL_PATH
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if excel_path.exists():
        try:
            wb = load_workbook(excel_path)
        except PermissionError:
            _log_message(
                f"ERROR: cannot open Excel for update: {excel_path}", level="ERROR"
            )
            return
        ws = wb.active
        ensure_excel_headers(ws)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Projects"
        ws.append(EXCEL_HEADERS)

    # Check if project already exists — prefer matching by unique project_number
    project_name = str(project_info.get("project_name", "") or "").strip()
    project_number = str(project_info.get("project_number", "") or "").strip()

    def _update_row(row_idx: int) -> None:
        ws.cell(row_idx, 1).value = project_info.get(
            "download_date", datetime.now().strftime("%Y-%m-%d")
        )
        ws.cell(row_idx, 8).value = project_info.get("tor_downloaded", "No")
        if "prelim_pricing" in project_info:
            ws.cell(row_idx, 9).value = project_info["prelim_pricing"]
        # Update keyword only if not set (avoid overwriting historical keyword)
        if not ws.cell(row_idx, 7).value:
            ws.cell(row_idx, 7).value = project_info.get("keyword", "")
        # Update search_name if provided
        if "search_name" in project_info:
            ws.cell(row_idx, 10).value = project_info["search_name"]
        if "tracking_status" in project_info:
            ws.cell(row_idx, 11).value = project_info["tracking_status"]
        if "closed_reason" in project_info:
            ws.cell(row_idx, 12).value = project_info["closed_reason"]
        if "artifact_bucket" in project_info:
            ws.cell(row_idx, 13).value = project_info["artifact_bucket"]
        # Backfill project_number if missing
        if project_number and not ws.cell(row_idx, 4).value:
            ws.cell(row_idx, 4).value = project_number

    if project_number:
        for row_idx in range(2, ws.max_row + 1):
            existing_num = ws.cell(row_idx, 4).value  # column D
            if existing_num and str(existing_num).strip() == project_number:
                _update_row(row_idx)
                wb.save(excel_path)
                return

    search_name = str(project_info.get("search_name", "") or "").strip()
    if search_name:
        for row_idx in range(2, ws.max_row + 1):
            existing_search_name = ws.cell(row_idx, 10).value  # column J
            existing_number = ws.cell(row_idx, 4).value  # column D
            if (
                existing_search_name
                and str(existing_search_name).strip() == search_name
                and (
                    not project_number
                    or not existing_number
                    or str(existing_number).strip() == project_number
                )
            ):
                _update_row(row_idx)
                wb.save(excel_path)
                return

    if project_name and project_number:
        for row_idx in range(2, ws.max_row + 1):
            existing_name = ws.cell(row_idx, 2).value  # column B
            existing_number = ws.cell(row_idx, 4).value  # column D
            existing_search_name = ws.cell(row_idx, 10).value  # column J
            if not (
                existing_name
                and str(existing_name).strip() == project_name
                and not str(existing_number or "").strip()
            ):
                continue
            if existing_search_name and search_name and str(existing_search_name).strip() != search_name:
                continue
            _update_row(row_idx)
            wb.save(excel_path)
            return

    if project_name and not project_number:
        for row_idx in range(2, ws.max_row + 1):
            existing_name = ws.cell(row_idx, 2).value  # column B
            existing_number = ws.cell(row_idx, 4).value  # column D
            if (
                existing_name
                and str(existing_name).strip() == project_name
                and not str(existing_number or "").strip()
            ):
                _update_row(row_idx)
                wb.save(excel_path)
                return

    # New project — append
    ws.append(
        [
            project_info.get("download_date", datetime.now().strftime("%Y-%m-%d")),
            project_info.get("project_name", ""),
            project_info.get("organization", ""),
            project_info.get("project_number", ""),
            project_info.get("budget", ""),
            project_info.get("proposal_submission_date", ""),
            project_info.get("keyword", ""),
            project_info.get("tor_downloaded", "No"),
            project_info.get("prelim_pricing", "No"),
            project_info.get("search_name", ""),
            project_info.get("tracking_status", ""),
            project_info.get("closed_reason", ""),
            project_info.get("artifact_bucket", ""),
        ]
    )
    wb.save(excel_path)


def load_existing_project_folder_map(excel_path: Path | None = None) -> dict[str, str]:
    """Return a mapping of project_number → sanitized folder name (project_name).

    This lets the crawler reuse the same project folder across runs even if the
    current scrape yields slightly different project_name text.
    """
    excel_path = Path(excel_path) if excel_path else EXCEL_PATH
    if not excel_path.exists():
        return {}

    try:
        wb = load_workbook(excel_path)
    except PermissionError:
        _log_message(f"ERROR: cannot read Excel folder map {excel_path}", level="ERROR")
        return {}

    ws = wb.active
    mapping: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        project_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        project_number = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        if not project_number or not project_name:
            continue
        mapping[project_number] = sanitize_dirname(project_name)
    return mapping


# ---------------------------------------------------------------------------
# Browser helpers
# ---------------------------------------------------------------------------


CHROME_PATH = env_get_str(
    "EGP_CHROME_PATH",
    "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
)
CDP_PORT = env_get_int("EGP_CDP_PORT", 9222, min_value=1, max_value=65535)


def launch_real_chrome() -> subprocess.Popen:
    """Launch the real Chrome browser with remote debugging enabled.

    Uses a dedicated profile directory so it doesn't interfere with
    the user's normal Chrome sessions.
    """
    BROWSER_PROFILE_DIR.mkdir(parents=True, exist_ok=True)

    proc = subprocess.Popen(
        [
            CHROME_PATH,
            f"--remote-debugging-port={CDP_PORT}",
            f"--user-data-dir={BROWSER_PROFILE_DIR}",
            "--no-first-run",
            "--no-default-browser-check",
            "--window-size=1280,900",
            "--disable-features=DownloadBubble",
        ],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    # Wait for Chrome to start CDP listener (Chrome can be slow to boot).
    if not wait_for_local_tcp_listen("127.0.0.1", CDP_PORT, timeout_seconds=15):
        print(
            f"WARNING: Chrome did not open CDP port {CDP_PORT} yet. "
            "If connection fails, close existing Chrome and retry."
        )
    return proc


def connect_playwright_to_chrome(pw):
    """Connect Playwright to the running Chrome instance via CDP."""
    if not wait_for_local_tcp_listen("127.0.0.1", CDP_PORT, timeout_seconds=15):
        raise RuntimeError(
            f"Chrome CDP port {CDP_PORT} is not reachable. "
            "Make sure Chrome launched with --remote-debugging-port, "
            "or set EGP_CDP_PORT to an unused port. "
            "If you have Chrome already open, fully quit it and retry."
        )
    browser = pw.chromium.connect_over_cdp(f"http://127.0.0.1:{CDP_PORT}")
    context = browser.contexts[0] if browser.contexts else browser.new_context()
    context.set_default_timeout(NAV_TIMEOUT)
    page = context.pages[0] if context.pages else context.new_page()
    return browser, page


def wait_for_local_tcp_listen(host: str, port: int, timeout_seconds: float) -> bool:
    """Return True once a TCP connect succeeds within timeout."""
    deadline = time.monotonic() + max(0.0, float(timeout_seconds))
    while True:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(0.5)
        try:
            sock.connect((host, int(port)))
            return True
        except OSError:
            pass
        finally:
            try:
                sock.close()
            except Exception:
                pass

        if time.monotonic() >= deadline:
            return False
        _REAL_SLEEP(0.15)


def safe_shutdown(
    *,
    browser=None,
    pw=None,
    chrome_proc: subprocess.Popen | None = None,
    ignore_sigint: bool = True,
) -> None:
    """Best-effort shutdown that still runs even after Ctrl+C.

    This aims to prevent leaving Playwright's driver process running if the user
    interrupts the crawler, and also cleans up the dedicated Chrome instance.
    """
    old_sigint = None
    if ignore_sigint:
        try:
            old_sigint = signal.getsignal(signal.SIGINT)
            signal.signal(signal.SIGINT, signal.SIG_IGN)
        except Exception:
            old_sigint = None

    try:
        if browser is not None:
            try:
                browser.close()
            except BaseException:
                pass

        if pw is not None:
            try:
                pw.stop()
            except BaseException:
                pass

        if chrome_proc is not None:
            try:
                chrome_proc.send_signal(signal.SIGTERM)
                chrome_proc.wait(timeout=5)
            except BaseException:
                try:
                    chrome_proc.kill()
                    chrome_proc.wait(timeout=5)
                except BaseException:
                    pass
    finally:
        if ignore_sigint and old_sigint is not None:
            try:
                signal.signal(signal.SIGINT, old_sigint)
            except Exception:
                pass


def wait_for_cloudflare(
    page,
    timeout_ms: int = CLOUDFLARE_TIMEOUT,
    reload_retries: int = CLOUDFLARE_RELOAD_RETRIES,
) -> bool:
    """Wait for Cloudflare Turnstile to pass.

    On the search page: checks if ค้นหา button is enabled (disabled until CF passes).
    On other pages (main page): waits for Cloudflare iframe to disappear or
    just waits a few seconds if no Cloudflare challenge is detected.
    """
    start = time.time()
    timeout_s = timeout_ms / 1000
    prompted = False

    while time.time() - start < timeout_s:
        # Check 1: Is the search button present and enabled? (search page)
        search_btn = page.query_selector(
            "button:has-text('ค้นหา'):not(:has-text('ค้นหาขั้นสูง'))"
        )
        if search_btn:
            is_disabled = search_btn.get_attribute("disabled")
            if is_disabled is None:  # No disabled attribute = enabled
                if prompted:
                    print("  Cloudflare: Passed!")
                return True

        # Check 2: No search button (main page or other page)
        # Wait for CF iframe to disappear, or timeout after brief wait
        if not search_btn:
            cf_iframe = page.query_selector("iframe[src*='challenges.cloudflare.com']")
            if not cf_iframe:
                # No Cloudflare challenge and no search button — page is ready
                return True

        if not prompted:
            print("  Cloudflare: Waiting for verification...", flush=True)
            prompted = True

        logged_sleep(2)

    if reload_retries > 0:
        print("  WARNING: Cloudflare timeout — reloading page and retrying")
        try:
            page.reload(wait_until="domcontentloaded", timeout=NAV_TIMEOUT)
        except Exception:
            try:
                page.goto(page.url, wait_until="domcontentloaded", timeout=NAV_TIMEOUT)
            except Exception:
                print("  WARNING: Cloudflare reload failed")
                print("  WARNING: Cloudflare timeout — continuing anyway")
                return False
        logged_sleep(3, "reload after Cloudflare timeout")
        return wait_for_cloudflare(
            page, timeout_ms=timeout_ms, reload_retries=reload_retries - 1
        )

    print("  WARNING: Cloudflare timeout — continuing anyway")
    return False


# ---------------------------------------------------------------------------
# Search & navigation
# ---------------------------------------------------------------------------


def _compact_visible_text(s: str | None) -> str:
    return re.sub(r"\s+", "", s or "")


RESULTS_TABLE_REQUIRED_HEADERS = [
    "ลำดับ",
    "หน่วยจัดซื้อ",
    "ชื่อโครงการ",
    "วงเงินงบประมาณ",
    "สถานะโครงการ",
    "ดูข้อมูล",
]


def status_matches_target(status_text: str) -> bool:
    """Normalize whitespace and compare against TARGET_STATUS."""
    return _TARGET_STATUS_COMPACT in _compact_visible_text(status_text)


def get_results_page_marker(page) -> dict[str, str | int]:
    """Capture a compact signature for the current results page."""
    rows = get_results_rows(page)[:3]
    row_sample_parts = []
    for row in rows:
        try:
            cells = row.query_selector_all("td")[:5]
            row_sample_parts.append(
                "|".join((cell.inner_text() or "").strip() for cell in cells)
            )
        except Exception:
            continue
    try:
        active = page.query_selector("li.page-item.active, li.active, .pagination .active")
        active_page = active.inner_text().strip() if active else ""
    except Exception:
        active_page = ""
    return {
        "active_page": active_page,
        "row_count": len(rows),
        "row_sample": " || ".join(row_sample_parts),
    }


def results_page_marker_changed(
    previous: dict[str, str | int], current: dict[str, str | int]
) -> bool:
    """Return True once pagination changes the active page or visible row sample."""
    return (
        str(previous.get("active_page", "") or "")
        != str(current.get("active_page", "") or "")
        or int(previous.get("row_count", -1) or -1)
        != int(current.get("row_count", -1) or -1)
        or str(previous.get("row_sample", "") or "")
        != str(current.get("row_sample", "") or "")
    )


def _table_matches_results_headers(table) -> bool:
    """Return True when a table looks like the main procurement results table."""
    header_selectors = ("thead th, thead td", "th", "tr:first-child th, tr:first-child td")
    headers: list[str] = []
    for selector in header_selectors:
        try:
            header_els = table.query_selector_all(selector)
        except Exception:
            header_els = []
        headers = [h.inner_text().strip() for h in header_els if h.inner_text().strip()]
        if headers:
            break

    if not headers:
        return False

    header_compact = [_compact_visible_text(h) for h in headers]
    return all(
        any(_compact_visible_text(required) in header for header in header_compact)
        for required in RESULTS_TABLE_REQUIRED_HEADERS
    )


def find_results_table(page):
    """Return the procurement search results table, if present."""
    for table in page.query_selector_all("table"):
        try:
            if _table_matches_results_headers(table):
                return table
        except Exception:
            continue
    return None


def get_results_rows(page) -> list:
    """Return rows from the procurement search results table only."""
    table = find_results_table(page)
    if not table:
        return []
    try:
        return table.query_selector_all("tbody tr")
    except Exception:
        return []


def find_search_input(page, search_btn):
    """Best-effort: locate the keyword search input near the ค้นหา button."""
    # Prefer an input inside the same form/container as the search button.
    try:
        handle = search_btn.evaluate_handle(
            """(btn) => {
                const root = btn.closest('form') || btn.closest('div') || document;
                const selectors = [
                    "input[placeholder*='ระบุ']",
                    "input[name*='keyword' i]",
                    "input[id*='keyword' i]",
                    "input[formcontrolname*='keyword' i]",
                    "input[type='text']",
                ];
                for (const sel of selectors) {
                    const el = root.querySelector(sel);
                    if (el) return el;
                }
                // Fallback: any visible text input on the page
                return document.querySelector("input[placeholder*='ระบุ'], input[type='text']");
            }"""
        )
        el = handle.as_element()
        if el:
            return el
    except Exception:
        pass

    # Fallback selector strategy (first visible candidate)
    candidates = page.query_selector_all(
        "input[placeholder*='ระบุ'], input[name*='keyword' i], input[id*='keyword' i], input[type='text']"
    )
    for c in candidates:
        try:
            if c.is_visible():
                return c
        except Exception:
            continue
    return page.wait_for_selector("input[type='text']", timeout=NAV_TIMEOUT)


def click_search_button(page, search_btn=None) -> None:
    """Click the primary search button with a DOM-query fallback for SPA re-renders."""
    try:
        clicked = page.evaluate(
            """() => {
                const buttons = Array.from(document.querySelectorAll('button'));
                for (const btn of buttons) {
                    const txt = (btn.innerText || '').trim();
                    if (
                        txt.includes('ค้นหา') &&
                        !txt.includes('ค้นหาขั้นสูง') &&
                        btn.offsetParent !== null &&
                        !btn.disabled
                    ) {
                        btn.click();
                        return true;
                    }
                }
                return false;
            }"""
        )
        if clicked:
            return
    except Exception:
        pass

    try:
        if search_btn is None:
            search_btn = page.wait_for_selector(
                "button:has-text('ค้นหา'):not(:has-text('ค้นหาขั้นสูง'))",
                timeout=NAV_TIMEOUT,
            )
        search_btn.click()
        return
    except Exception:
        fallback_btn = page.wait_for_selector(
            "button:has-text('ค้นหา'):not(:has-text('ค้นหาขั้นสูง'))",
            timeout=NAV_TIMEOUT,
        )
        page.evaluate("(el) => el.click()", fallback_btn)


def is_no_results_page(page) -> bool:
    """Return True only when the results table is empty and shows an empty-state message."""
    try:
        table = find_results_table(page)
        if not table:
            return False
        rows = get_results_rows(page)
        if rows:
            return False
        text = re.sub(r"\s+", " ", table.inner_text() or "").strip()
        return "ไม่พบข้อมูล" in text or "จำนวนโครงการที่พบ : 0" in text
    except Exception:
        return False


def wait_for_results_ready(page) -> None:
    """Wait until results rows appear or an empty-state is reliably shown."""
    page.wait_for_selector("table", state="attached", timeout=NAV_TIMEOUT)

    try:
        page.wait_for_function(
            """() => {
                const table = document.querySelector('table');
                if (!table) return false;
                const rows = table.querySelectorAll('tbody tr');
                if (rows.length > 0) return true;
                const txt = (table.innerText || '').replace(/\\s+/g, ' ').trim();
                return txt.includes('ไม่พบข้อมูล') || txt.includes('จำนวนโครงการที่พบ : 0');
            }""",
            timeout=NAV_TIMEOUT,
        )
    except Exception:
        pass

    # Guard against premature "ไม่พบข้อมูล" placeholder while the SPA is still loading.
    for _ in range(3):
        try:
            if get_results_rows(page):
                return
        except Exception:
            pass
        if is_no_results_page(page):
            logged_sleep(1.0, "confirm empty results")
        else:
            logged_sleep(1.0, "wait results render")


def wait_for_results_page_change(
    page, previous_marker: dict[str, str | int], timeout_ms: int = NAV_TIMEOUT
) -> bool:
    """Wait for the results table to move to a different page or row sample."""
    deadline = time.monotonic() + max(1.0, timeout_ms / 1000)
    while time.monotonic() < deadline:
        wait_for_results_ready(page)
        current_marker = get_results_page_marker(page)
        if results_page_marker_changed(previous_marker, current_marker):
            return True
        if is_no_results_page(page):
            return True
        logged_sleep(0.5, "wait page change")
    current_marker = get_results_page_marker(page)
    return results_page_marker_changed(previous_marker, current_marker) or is_no_results_page(page)


def search_keyword(
    page, keyword: str, search_page_retries: int = SEARCH_PAGE_RECOVERY_RETRIES
) -> None:
    """Enter keyword on the search page and submit. Assumes page is already on search URL."""
    # Wait for Cloudflare to pass (ค้นหา button is disabled until it does)
    cloudflare_ok = wait_for_cloudflare(page)
    if not cloudflare_ok and search_page_retries > 0:
        print("  Cloudflare still blocked — reopening search page and retrying keyword")
        page.goto(SEARCH_URL, wait_until="domcontentloaded", timeout=NAV_TIMEOUT)
        logged_sleep(3, "reopen search page after Cloudflare timeout")
        search_keyword(page, keyword, search_page_retries=search_page_retries - 1)
        return

    search_btn = page.query_selector(
        "button:has-text('ค้นหา'):not(:has-text('ค้นหาขั้นสูง'))"
    )
    if not search_btn:
        search_btn = page.wait_for_selector(
            "button:has-text('ค้นหา'):not(:has-text('ค้นหาขั้นสูง'))",
            timeout=NAV_TIMEOUT,
        )

    search_input = find_search_input(page, search_btn)

    # Clear existing text and type keyword
    search_input.click()
    search_input.fill("")
    search_input.fill(keyword)
    logged_sleep(0.5)

    # Click the search button (already confirmed enabled by wait_for_cloudflare)
    click_search_button(page, search_btn)

    wait_for_results_ready(page)

    row_count = -1
    last = None
    stable_polls = 0
    # Give the SPA a moment to finish rendering/paginating rows.
    for _ in range(10):
        try:
            current_count = len(get_results_rows(page))
        except Exception:
            current_count = -1
        row_count = max(row_count, current_count)
        if current_count == last:
            stable_polls += 1
        else:
            stable_polls = 0
        if current_count > 0 and stable_polls >= 2:
            break
        last = current_count
        logged_sleep(0.5, "results stabilize")
    print(f"  Searched for: {keyword} (rows: {row_count}, url: {page.url})")


def restore_results_page(page, keyword: str, target_page_num: int) -> None:
    """Re-run the keyword search and advance back to the requested results page."""
    search_keyword(page, keyword)
    current_page = 1
    while current_page < max(target_page_num, 1):
        dismiss_modal(page)
        previous_marker = get_results_page_marker(page)
        next_btn = page.query_selector(NEXT_PAGE_SELECTOR)
        if not (next_btn and next_btn.is_visible()):
            break
        try:
            page.evaluate("(el) => el.click()", next_btn)
        except Exception:
            next_btn.click(timeout=10_000)
        logged_sleep(3, f"restore results page {current_page + 1}")
        if not wait_for_results_page_change(page, previous_marker):
            break
        current_page += 1


def build_results_debug_snapshot(page, sample_limit: int = 3) -> dict[str, object]:
    """Capture a compact diagnostic snapshot of the current search results view."""
    snapshot: dict[str, object] = {
        "url": getattr(page, "url", ""),
        "active_page": "",
        "results_headers": [],
        "results_row_count": 0,
        "results_row_samples": [],
        "table_count": 0,
        "body_snippet": "",
    }

    try:
        tables = page.query_selector_all("table")
    except Exception:
        tables = []
    snapshot["table_count"] = len(tables)

    table = find_results_table(page)
    if table:
        header_selectors = ("thead th, thead td", "th", "tr:first-child th, tr:first-child td")
        headers: list[str] = []
        for selector in header_selectors:
            try:
                header_els = table.query_selector_all(selector)
            except Exception:
                header_els = []
            headers = [h.inner_text().strip() for h in header_els if h.inner_text().strip()]
            if headers:
                break
        snapshot["results_headers"] = headers

        rows = get_results_rows(page)
        snapshot["results_row_count"] = len(rows)
        samples: list[list[str]] = []
        for row in rows[:sample_limit]:
            try:
                samples.append(
                    [cell.inner_text().strip() for cell in row.query_selector_all("td")]
                )
            except Exception:
                continue
        snapshot["results_row_samples"] = samples

    try:
        active = page.query_selector("li.page-item.active, li.active, .pagination .active")
        snapshot["active_page"] = active.inner_text().strip() if active else ""
    except Exception:
        pass

    try:
        body_text = page.inner_text("body")
        body_lines = _split_visible_lines(body_text)
        snapshot["body_snippet"] = " | ".join(body_lines[:10])
    except Exception:
        pass

    return snapshot


def log_results_debug_snapshot(page, keyword: str, reason: str) -> None:
    """Print a compact diagnostic snapshot for unexpected search-result states."""
    snapshot = build_results_debug_snapshot(page)
    print(
        f"    DEBUG [{reason}] keyword={keyword} active_page={snapshot['active_page'] or '-'} "
        f"tables={snapshot['table_count']} results_rows={snapshot['results_row_count']} "
        f"url={snapshot['url']}"
    )
    headers = snapshot.get("results_headers") or []
    if headers:
        print(f"    DEBUG headers: {' | '.join(str(h) for h in headers)}")
    for idx, row in enumerate(snapshot.get("results_row_samples") or [], start=1):
        print(f"    DEBUG row{idx}: {' | '.join(str(cell) for cell in row)}")
    body_snippet = str(snapshot.get("body_snippet") or "")
    if body_snippet:
        print(f"    DEBUG body: {body_snippet[:500]}")


def _split_visible_lines(text: str) -> list[str]:
    return [line.strip() for line in re.split(r"[\r\n]+", text or "") if line.strip()]


def clear_search(page) -> None:
    """Click the clear button to reset search filters."""
    try:
        clear_btn = page.wait_for_selector(
            "button:has-text('ล้างตัวเลือก')",
            timeout=10_000,
        )
        clear_btn.click()
        logged_sleep(1)
    except PlaywrightTimeout:
        # If no clear button, navigate to search URL
        page.goto(SEARCH_URL, wait_until="domcontentloaded", timeout=NAV_TIMEOUT)
        logged_sleep(3)
        wait_for_cloudflare(page)


def collect_eligible_project_links(page) -> list[dict]:
    """
    Scan results table across all pages. Collect projects where
    สถานะโครงการ = 'หนังสือเชิญชวน/ประกาศเชิญชวน'.
    Returns list of dicts with 'row_index' and basic info for identification.
    """
    eligible = []
    page_num = 1

    while True:
        print(f"    Scanning results page {page_num}...")
        rows = get_results_rows(page)

        for row in rows:
            cells = row.query_selector_all("td")
            if len(cells) < 6:
                continue

            status_text = cells[4].inner_text().strip()
            if not status_matches_target(status_text):
                continue

            project_name = cells[2].inner_text().strip()
            org = cells[1].inner_text().strip()

            # Get the ดูข้อมูล link/button in the last column
            view_btn = cells[5].query_selector("a, button, [role='button'], svg, i")

            eligible.append(
                {
                    "project_name_preview": project_name[:80],
                    "organization_preview": org,
                    "view_element": view_btn,
                    "row": row,
                }
            )

        # Check for next page
        next_btn = page.query_selector(
            "button[aria-label='next'], "
            "a:has-text('»'), "
            "li.next:not(.disabled) a, "
            "button:has-text('ถัดไป'), "
            "[class*='next']:not([class*='disabled'])"
        )

        if next_btn and next_btn.is_visible() and next_btn.is_enabled():
            previous_marker = get_results_page_marker(page)
            next_btn.click()
            logged_sleep(2)
            if not wait_for_results_page_change(page, previous_marker):
                break
            page_num += 1
        else:
            break

    print(f"    Found {len(eligible)} eligible projects")
    return eligible


def navigate_to_project_by_row(page, row_index: int) -> bool:
    """
    Re-query the table and click ดูข้อมูล for the row at given index.
    Returns True if navigation succeeds.
    """
    rows = get_results_rows(page)

    # Find eligible rows again and click the one at row_index
    eligible_idx = 0
    for row in rows:
        cells = row.query_selector_all("td")
        if len(cells) < 6:
            continue
        status_text = cells[4].inner_text().strip()
        if TARGET_STATUS not in status_text:
            continue

        if eligible_idx == row_index:
            view_btn = cells[5].query_selector("a, button, [role='button'], svg, i")
            if view_btn:
                view_btn.click()
                logged_sleep(2)
                return True
            # Try clicking the cell itself
            cells[5].click()
            logged_sleep(2)
            return True

        eligible_idx += 1

    return False


def check_has_preliminary_pricing(page) -> bool:
    """Check if project page has สรุปข้อมูลการเสนอราคาเบื้องต้น."""
    content = page.content()
    return SKIP_TEXT in content


def check_announcement_stale(page, max_days: int = STALE_ANNOUNCEMENT_DAYS) -> bool:
    """Check if the invitation announcement date is older than max_days.

    Looks for rows containing หนังสือเชิญชวน or ประกาศเชิญชวน in the
    เอกสาร/ประกาศที่เกี่ยวข้อง table, extracts วันที่ประกาศ, and checks
    if it's more than max_days old.

    Returns True if stale (should delete), False otherwise.
    """
    tables = page.query_selector_all("table")
    for table in tables:
        rows = table.query_selector_all("tbody tr")
        for row in rows:
            cells = row.query_selector_all("td")
            if len(cells) < 3:
                continue
            row_text = " ".join(c.inner_text().strip() for c in cells)
            if "หนังสือเชิญชวน" in row_text or "ประกาศเชิญชวน" in row_text:
                # Extract the date from the วันที่ประกาศ column
                for cell in cells:
                    cell_text = cell.inner_text().strip()
                    parsed = parse_buddhist_date(cell_text)
                    if parsed:
                        age_days = (date.today() - parsed).days
                        if age_days > max_days:
                            print(
                                f"      Announcement date: {cell_text} ({age_days} days ago) — stale"
                            )
                            return True
                        else:
                            print(
                                f"      Announcement date: {cell_text} ({age_days} days ago) — still fresh"
                            )
                            return False
    return False  # No announcement date found — don't delete


def extract_project_info(page) -> dict:
    """Extract project details from the project info page."""
    info = {
        "download_date": datetime.now().strftime("%Y-%m-%d"),
        "project_name": "",
        "organization": "",
        "project_number": "",
        "budget": "",
        "proposal_submission_date": "",
    }

    # Try to extract from the page structure
    # The page has a table-like layout with label-value pairs
    body_text = page.inner_text("body")

    # Extract project name — look for ชื่อโครงการ label
    name_match = re.search(
        r"ชื่อโครงการ\s*[:\s]\s*(.+?)(?:\n|เลขที่)", body_text, re.DOTALL
    )
    if name_match:
        info["project_name"] = name_match.group(1).strip()

    # Extract organization — look for หน่วยจัดซื้อ or หน่วยงาน
    org_match = re.search(
        r"(?:หน่วยจัดซื้อ|หน่วยงาน)\s*[:\s]\s*(.+?)(?:\n|ชื่อโครงการ|วิธี)", body_text, re.DOTALL
    )
    if org_match:
        info["organization"] = org_match.group(1).strip()

    # Extract project number — look for เลขที่โครงการ
    num_match = re.search(r"เลขที่โครงการ\s*[:\s]\s*(\S+)", body_text)
    if num_match:
        info["project_number"] = num_match.group(1).strip()

    # Extract budget — look for วงเงินงบประมาณ value
    budget_match = re.search(r"วงเงินงบประมาณ\s*\n?\s*([\d,]+\.?\d*)", body_text)
    if budget_match:
        info["budget"] = budget_match.group(1).strip()

    # Extract proposal submission date — look for วันที่ยื่นข้อเสนอ or ยื่นซอง
    date_match = re.search(
        r"(?:วันที่ยื่นข้อเสนอ|ยื่นซอง|วันยื่นข้อเสนอ|สิ้นสุดยื่นข้อเสนอ)\s*[:\s]\s*(\d{2}/\d{2}/\d{4})",
        body_text,
    )
    if date_match:
        info["proposal_submission_date"] = date_match.group(1).strip()
    else:
        # Try another pattern
        date_match2 = re.search(
            r"(?:ยื่นข้อเสนอ|submit).+?(\d{2}/\d{2}/\d{4})",
            body_text,
            re.IGNORECASE,
        )
        if date_match2:
            info["proposal_submission_date"] = date_match2.group(1).strip()

    return info


# ---------------------------------------------------------------------------
# Document downloading
# ---------------------------------------------------------------------------


def download_project_documents(page, project_dir: Path) -> ProjectDocumentSummary:
    """Download target documents from the project info page.

    Downloads draft TOR, final TOR, invitation, and price-announcement artifacts.
    Draft TORs are saved for public-hearing monitoring, but the return value only
    marks completion once the final invitation-stage TOR is available.

    Re-queries the DOM before each document to handle page state changes
    after navigating to/from sub-pages.

    Returns a compact artifact summary for downstream state/export decisions.
    """
    project_dir.mkdir(parents=True, exist_ok=True)
    saved_labels: list[str] = []

    for target_doc in DOCS_TO_DOWNLOAD:
        try:
            saved_labels.extend(_download_one_document(page, target_doc, project_dir))
        except Exception as e:
            print(f"      ERROR downloading {target_doc}: {e}")
            # Skip this file, continue to next document type

    artifact_bucket = derive_artifact_bucket(labels=saved_labels)
    if artifact_bucket is ArtifactBucket.DRAFT_PLUS_PRICING:
        print("      Draft TOR downloaded; final TOR not available yet")
    return ProjectDocumentSummary(
        saved_labels=tuple(saved_labels),
        artifact_bucket=artifact_bucket,
    )


def _download_one_document(page, target_doc: str, project_dir: Path) -> list[str]:
    """Find and download a single document type from the project info page.

    Only searches tables that have a ดูข้อมูล or ดาวน์โหลด column header —
    some tables (e.g. Table 3 with ประกาศเชิญชวน) just list document names
    with dates but have no download functionality.

    Returns:
        Saved source labels for the downloaded artifacts.
    """
    dismiss_modal(page)
    logged_sleep(0.5)
    is_draft_tor_target = target_doc == "ร่างเอกสารประกวดราคา"
    is_final_tor_target = target_doc == "เอกสารประกวดราคา"

    # Find tables that have a download column (ดูข้อมูล or ดาวน์โหลด in headers)
    tables = page.query_selector_all("table")
    downloadable_rows = []
    for table in tables:
        header_els = table.query_selector_all("th")
        header_combined = " ".join(h.inner_text().strip() for h in header_els)
        if "ดูข้อมูล" not in header_combined and "ดาวน์โหลด" not in header_combined:
            continue
        # This table has a download column — collect its rows
        for row in table.query_selector_all("tbody tr"):
            downloadable_rows.append(row)

    for row in downloadable_rows:
        cells = row.query_selector_all("td")
        if len(cells) < 3:
            continue

        # Check if this row matches the target document
        doc_name = ""
        for cell in cells:
            text = cell.inner_text().strip()
            if target_doc in text:
                doc_name = text
                break
            if is_draft_tor_target and is_draft_tor_doc_label(text):
                doc_name = text
                break
            if is_final_tor_target and is_final_tor_doc_label(text):
                doc_name = text
                break

        if not doc_name:
            continue

        print(f"      Downloading: {doc_name}")

        # Find the clickable ดูข้อมูล element in the last cell
        last_cell = cells[-1]
        clickable = (
            last_cell.query_selector(
                "a[href], a[onclick], button:not([disabled]), [role='button']"
            )
            or last_cell.query_selector("a, button, [role='button']")
            or last_cell
        )

        if is_draft_tor_doc_label(doc_name):
            dismiss_modal(page)
            return _handle_subpage_download(
                page,
                clickable,
                project_dir,
                include_label=lambda label: is_tor_file(label),
            )

        saved_name = _handle_direct_or_page_download(page, clickable, project_dir, doc_name)
        if saved_name:
            if target_doc == "ประกาศเชิญชวน":
                return [doc_name]
            if is_final_tor_target and is_tor_file(saved_name):
                return [doc_name]
            if not is_final_tor_target:
                return [doc_name]
            return []

        if target_doc == "ประกาศเชิญชวน":
            return _download_documents_from_current_view(
                page,
                project_dir,
                include_label=lambda label: "ประกาศเชิญชวน" in label
                or is_final_tor_doc_label(label),
            )

        if is_final_tor_target:
            return _download_documents_from_current_view(
                page,
                project_dir,
                include_label=is_final_tor_doc_label,
            )

        return []

    return []  # Document not found in any downloadable table


def _save_download_to_project(
    download, project_dir: Path, filename: str | None = None
) -> str:
    """Save a Playwright download to temp dir, then move to project dir in OneDrive.

    Downloading directly to OneDrive can trigger native macOS OneDrive popups
    that block the browser. We save to /tmp first, then move.
    """
    TEMP_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    project_dir.mkdir(parents=True, exist_ok=True)

    fname = filename or download.suggested_filename
    tmp_path = TEMP_DOWNLOAD_DIR / fname
    download.save_as(tmp_path)

    final_path = project_dir / fname
    try:
        shutil.move(str(tmp_path), str(final_path))
        print(f"      Saved: {final_path.name}")
        return final_path.name
    except PermissionError:
        fallback_final = _mirror_to_local_fallback(project_dir) / fname
        fallback_final.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(tmp_path), str(fallback_final))
        _log_message(
            f"Download move blocked for {final_path}; saved to fallback {fallback_final}",
            level="WARN",
        )
        print(f"      Saved (fallback): {fallback_final.name}")
        return fallback_final.name


def _save_bytes_to_project(project_dir: Path, filename: str, data: bytes) -> str | None:
    """Save raw bytes via temp dir, then move to project dir in OneDrive."""
    if not filename:
        return None
    TEMP_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    project_dir.mkdir(parents=True, exist_ok=True)

    tmp_path = TEMP_DOWNLOAD_DIR / filename
    tmp_path.write_bytes(data)

    final_path = project_dir / filename
    try:
        shutil.move(str(tmp_path), str(final_path))
        print(f"      Saved: {final_path.name}")
        return final_path.name
    except PermissionError:
        fallback_final = _mirror_to_local_fallback(project_dir) / filename
        fallback_final.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(tmp_path), str(fallback_final))
        _log_message(
            f"Download move blocked for {final_path}; saved to fallback {fallback_final}",
            level="WARN",
        )
        print(f"      Saved (fallback): {fallback_final.name}")
        return fallback_final.name


def _infer_document_url_from_page(page) -> str | None:
    """Best-effort inference of the underlying file URL from a viewer page."""
    try:
        viewer_url = page.url
    except Exception:
        viewer_url = ""

    candidate = extract_document_url_from_viewer_url(viewer_url)
    if candidate:
        return candidate

    # Direct file URL (PDF/ZIP) shown in the tab.
    if viewer_url.startswith(("http://", "https://")):
        lowered = viewer_url.lower()
        if any(x in lowered for x in (".pdf", ".zip", "download", "dl=")):
            return viewer_url

    try:
        src = page.evaluate(
            """() => {
                const el = document.querySelector('embed[src], iframe[src], object[data]');
                if (!el) return null;
                return el.getAttribute('src') || el.getAttribute('data');
            }"""
        )
        if src:
            extracted = extract_document_url_from_viewer_url(str(src))
            resolved = extracted or str(src)
            if resolved.startswith(("http://", "https://")):
                return resolved
    except Exception:
        pass

    # Last resort: scan DOM links for a likely downloadable URL.
    try:
        urls = page.evaluate(
            """() => {
                const out = [];
                const add = (u) => {
                    if (!u) return;
                    try { out.push(new URL(u, window.location.href).toString()); } catch (e) {}
                };
                document.querySelectorAll('a[href]').forEach(a => add(a.getAttribute('href')));
                document.querySelectorAll('embed[src], iframe[src], object[data]').forEach(el => {
                    add(el.getAttribute('src') || el.getAttribute('data'));
                });
                return out.slice(0, 200);
            }"""
        )
        for u in urls or []:
            raw = str(u)
            extracted = extract_document_url_from_viewer_url(raw)
            resolved = extracted or raw
            if not resolved.startswith(("http://", "https://")):
                continue
            lowered = resolved.lower()
            if any(x in lowered for x in (".pdf", ".zip", "download", "file=", "dl=")):
                return resolved
    except Exception:
        pass

    return None


def _save_via_request(
    page, project_dir: Path, doc_name: str, fallback_url: str | None = None
) -> str | None:
    """Fetch the underlying file URL via Playwright request and save bytes."""
    url = _infer_document_url_from_page(page)
    if not url:
        try:
            base = page.url
        except Exception:
            base = None
        url = resolve_http_url(fallback_url, base_url=base)
    if not url:
        try:
            viewer_url = page.url
        except Exception:
            viewer_url = ""
        print(
            f"      Viewer fetch: could not infer file URL for {doc_name} "
            f"(viewer: {viewer_url}; clicked: {fallback_url or ''})"
        )
        return None

    if not is_allowed_download_url(url):
        try:
            viewer_url = page.url
        except Exception:
            viewer_url = ""
        print(
            "      Viewer fetch blocked (URL outside allowlist): "
            f"{url} (viewer: {viewer_url}; clicked: {fallback_url or ''})"
        )
        return None

    print(f"      Viewer fetch URL: {url}")
    try:
        resp = page.request.get(url, timeout=SUBPAGE_DOWNLOAD_TIMEOUT)
    except Exception as e:
        print(f"      Viewer fetch failed for {doc_name}: {e}")
        return None

    if not resp.ok:
        ct = resp.headers.get("content-type") if hasattr(resp, "headers") else None
        print(
            f"      Viewer fetch failed for {doc_name}: HTTP {resp.status} (content-type: {ct})"
        )
        return None

    try:
        data = resp.body()
    except Exception as e:
        print(f"      Viewer fetch failed reading body for {doc_name}: {e}")
        return None

    cd = resp.headers.get("content-disposition") if hasattr(resp, "headers") else None
    guessed = filename_from_content_disposition(cd)
    ct = resp.headers.get("content-type") if hasattr(resp, "headers") else None
    sniffed_ext = sniff_extension_from_bytes(data)

    if looks_like_html_bytes(data) or (ct and "text/html" in ct.lower()):
        snippet = data.lstrip()[:120].decode("utf-8", errors="replace")
        print(
            f"      Viewer fetch rejected for {doc_name}: HTML/error response "
            f"(content-type: {ct}; head: {snippet!r})"
        )
        return None

    ext = (
        sniffed_ext
        or guess_extension_from_content_type(ct)
        or (Path(urlparse(url).path).suffix or None)
    )
    if not ext:
        print(
            f"      Viewer fetch rejected for {doc_name}: unknown file type (content-type: {ct})"
        )
        return None

    if guessed:
        filename = sanitize_filename_preserve_suffix(guessed, max_len=100)
        if sniffed_ext and not filename.lower().endswith(sniffed_ext):
            filename = build_safe_filename(
                Path(filename).stem, sniffed_ext, max_len=100
            )
    else:
        filename = build_safe_filename(doc_name, ext, max_len=100)

    return _save_bytes_to_project(project_dir, filename, data)


def _cancel_pending_downloads(page) -> None:
    """Cancel all in-progress Chrome downloads to prevent memory accumulation.

    When expect_download times out, Chrome still has the download active.
    Accumulated zombie downloads eventually crash Chrome's renderer.
    """
    try:
        page.evaluate("""() => {
            // Navigate away briefly cancels pending downloads in the renderer
            window.stop();
        }""")
    except Exception:
        pass


def clear_site_error_toast(page) -> bool:
    """Close the red e-GP runtime error toast if visible.

    This toast can appear intermittently and block interaction flow until the
    close button is clicked.
    """
    try:
        closed = page.evaluate(
            """() => {
                const compact = (s) => (s || '').replace(/\\s+/g, '');
                const isVisible = (el) => {
                    if (!el) return false;
                    const style = window.getComputedStyle(el);
                    return style.display !== 'none'
                        && style.visibility !== 'hidden'
                        && style.opacity !== '0'
                        && (el.offsetParent !== null || style.position === 'fixed');
                };

                const matchesToast = (el) => {
                    const txt = compact(el.textContent || '');
                    return txt.includes('ระบบเกิดข้อผิดพลาด') && txt.includes('กรุณาตรวจสอบ');
                };

                const candidates = Array.from(
                    document.querySelectorAll('[role="alert"], .toast, .alert, .toast-error, .swal2-popup')
                ).filter(isVisible);

                const toast = candidates.find(matchesToast);
                if (!toast) return false;

                const closeSelectors = [
                    '.toast-close-button',
                    '.close',
                    '[aria-label*="close" i]',
                    '[aria-label*="ปิด"]',
                    'button',
                ];

                for (const sel of closeSelectors) {
                    const btn = toast.querySelector(sel);
                    if (btn && isVisible(btn)) {
                        btn.click();
                        return true;
                    }
                }

                // Fallback: click a visible button-like element with "x"
                const maybe = Array.from(toast.querySelectorAll('span, i, div'))
                    .find(el => {
                        const t = (el.textContent || '').trim();
                        return isVisible(el) && /^(x|X|×|✕)$/.test(t);
                    });
                if (maybe) {
                    maybe.click();
                    return true;
                }
                return false;
            }"""
        )
        if closed:
            print("      Closed site error toast and retrying step")
            logged_sleep(0.3)
            page.keyboard.press("Escape")
            logged_sleep(0.2)
        return bool(closed)
    except Exception:
        return False


def run_with_toast_recovery(
    page, action, label: str, retries: int = TOAST_RECOVERY_RETRIES
):
    """Run an action, closing site error toasts and retrying on timeout."""
    clear_site_error_toast(page)
    for attempt in range(retries + 1):
        try:
            return action()
        except PlaywrightTimeout:
            had_toast = clear_site_error_toast(page)
            _cancel_pending_downloads(page)
            if attempt < retries:
                reason = "closed site error toast" if had_toast else "no download event"
                print(
                    f"      {label}: timeout ({reason}), retrying click ({attempt + 1}/{retries})"
                )
                logged_sleep(0.8)
                continue
            raise


def _handle_direct_or_page_download(
    page, btn, project_dir: Path, doc_name: str
) -> str | None:
    """Handle ประกาศเชิญชวน / ประกาศราคากลาง — two possible behaviors:

    Case 1: Click ดูข้อมูล → file downloads automatically (may show Chrome
            download popup or macOS popup or nothing).
    Case 2: Click ดูข้อมูล → opens new page showing file content →
            need to save from that page, then go back.
    """
    dismiss_modal(page)
    clear_site_error_toast(page)
    url_before = page.url
    pages_before = list(page.context.pages)
    try:
        click_meta = btn.evaluate(
            """(el) => {
                if (!el) return { href: null, onclick: null, tag: null };
                const a = el.closest && el.closest('a[href]') ? el.closest('a[href]') : null;
                const href = (a && a.href) || el.href || el.getAttribute('href') || null;
                const onclick = el.getAttribute ? el.getAttribute('onclick') : null;
                const tag = el.tagName ? String(el.tagName).toLowerCase() : null;
                return { href, onclick, tag };
            }"""
        )
    except Exception:
        click_meta = {"href": None, "onclick": None, "tag": None}

    clicked_href = (click_meta or {}).get("href")
    clicked_onclick = (click_meta or {}).get("onclick")
    if clicked_href:
        print(f"      Click target href: {clicked_href}")
    elif clicked_onclick:
        print(f"      Click target onclick: {clicked_onclick}")
    else:
        print(f"      Click target tag: {(click_meta or {}).get('tag')}")

    # Try to catch an automatic download
    try:

        def _click_and_wait_download():
            with page.expect_download(timeout=DOWNLOAD_EVENT_TIMEOUT) as download_info:
                page.evaluate("(el) => el.click()", btn)
            return download_info.value

        download = run_with_toast_recovery(
            page,
            _click_and_wait_download,
            "Direct download",
            retries=DOWNLOAD_CLICK_RETRIES,
        )
        ext = Path(download.suggested_filename).suffix or ".pdf"
        saved_name = _save_download_to_project(
            download, project_dir, build_safe_filename(doc_name, ext)
        )
        # Dismiss any Chrome download popup that appeared
        dismiss_modal(page)
        page.keyboard.press("Escape")
        return saved_name
    except PlaywrightTimeout:
        _cancel_pending_downloads(page)

    # No auto-download — check if page navigated to show file content
    logged_sleep(1)
    url_after = page.url
    if url_after != url_before:
        print("      Page navigated to content view, saving...")
        return _save_from_content_page(page, project_dir, doc_name)

    # Check if a new tab/page opened
    pages_after = list(page.context.pages)
    new_pages = [p for p in pages_after if p not in pages_before]
    if new_pages:
        new_page = new_pages[-1]
        print("      New tab opened, saving...")
        saved_name = _save_from_new_tab(
            new_page, project_dir, doc_name, fallback_url=clicked_href
        )
        try:
            new_page.close()
        except Exception as e:
            print(f"      WARN: failed to close viewer tab for {doc_name}: {e}")
        return saved_name

    # Some announcement links are popups (onclick) and may be blocked when clicked via JS.
    try:
        inferred = extract_url_from_onclick(clicked_onclick, base_url=page.url)
    except Exception:
        inferred = None
    if inferred and is_allowed_download_url(inferred):
        print(f"      Opening inferred viewer URL: {inferred}")
        try:
            new_page = page.context.new_page()
            new_page.goto(inferred, wait_until="domcontentloaded", timeout=NAV_TIMEOUT)
            saved_name = _save_from_new_tab(
                new_page, project_dir, doc_name, fallback_url=inferred
            )
            try:
                new_page.close()
            except Exception:
                pass
            return saved_name
        except Exception as e:
            print(f"      Could not open inferred viewer URL for {doc_name}: {e}")

    print(f"      Could not download {doc_name} (no download, no new page)")
    return None


def _save_from_content_page(page, project_dir: Path, doc_name: str) -> str | None:
    """Save file from a content page (e.g. PDF shown in browser), then go back."""
    project_dir.mkdir(parents=True, exist_ok=True)

    saved_name: str | None = None
    try:
        if is_show_htmlfile_url(page.url):
            print("      Content view is ShowHTMLFile; saving as PDF/HTML...")
            saved_name = save_show_htmlfile_as_file(
                page, project_dir, doc_name, prefer_pdf=True
            )
    except Exception:
        saved_name = None

    if saved_name:
        page.go_back()
        logged_sleep(2)
        return saved_name

    # Try Ctrl+S to trigger browser save dialog → expect_download
    try:

        def _save_with_ctrl_s():
            with page.expect_download(timeout=DOWNLOAD_TIMEOUT) as download_info:
                page.keyboard.press("Control+s")
            return download_info.value

        download = run_with_toast_recovery(
            page, _save_with_ctrl_s, "Content save (Ctrl+S)"
        )
        ext = Path(download.suggested_filename).suffix or ".pdf"
        saved_name = _save_download_to_project(
            download, project_dir, build_safe_filename(doc_name, ext)
        )
    except PlaywrightTimeout:
        # Ctrl+S didn't work — try to download the URL directly
        try:
            current_url = page.url
            escaped_url = json.dumps(current_url)
            escaped_name = json.dumps(sanitize_filename(doc_name)[:80] or "file")

            def _save_from_url():
                with page.expect_download(timeout=DOWNLOAD_TIMEOUT) as download_info:
                    page.evaluate(f"""() => {{
                        const a = document.createElement('a');
                        a.href = {escaped_url};
                        a.download = {escaped_name};
                        document.body.appendChild(a);
                        a.click();
                        a.remove();
                    }}""")
                return download_info.value

            download = run_with_toast_recovery(
                page, _save_from_url, "Content save (URL)"
            )
            ext = Path(download.suggested_filename).suffix or ".pdf"
            saved_name = _save_download_to_project(
                download, project_dir, build_safe_filename(doc_name, ext)
            )
        except (PlaywrightTimeout, Exception):
            print(f"      Could not save from content page for {doc_name}")
            _cancel_pending_downloads(page)

    # Fallback: fetch bytes via Playwright request (works for viewer pages)
    if not saved_name:
        try:
            saved_name = _save_via_request(page, project_dir, doc_name)
        except Exception:
            saved_name = None

    # Go back to project info page
    page.go_back()
    logged_sleep(2)
    return saved_name


def _save_from_new_tab(
    new_page, project_dir: Path, doc_name: str, fallback_url: str | None = None
) -> str | None:
    """Save file from a newly opened tab, then close it."""
    project_dir.mkdir(parents=True, exist_ok=True)

    try:
        new_page.wait_for_load_state("domcontentloaded", timeout=10_000)
    except Exception:
        pass
    # Some sites briefly open a blank new tab before navigating to a viewer URL.
    for _ in range(10):
        try:
            u = new_page.url
        except Exception:
            u = ""
        if u and not u.startswith(("chrome://new-tab-page", "about:blank")):
            break
        logged_sleep(0.5, "wait viewer tab navigate")
    try:
        print(f"      Viewer tab URL: {new_page.url}")
    except Exception:
        pass

    try:
        if is_show_htmlfile_url(new_page.url):
            print("      Viewer tab is ShowHTMLFile; saving as PDF/HTML...")
            return save_show_htmlfile_as_file(
                new_page, project_dir, doc_name, prefer_pdf=True
            )
    except Exception:
        pass

    try:

        def _save_from_tab():
            with new_page.expect_download(timeout=DOWNLOAD_TIMEOUT) as download_info:
                new_page.keyboard.press("Control+s")
            return download_info.value

        download = run_with_toast_recovery(new_page, _save_from_tab, "New-tab save")
        ext = Path(download.suggested_filename).suffix or ".pdf"
        return _save_download_to_project(
            download, project_dir, build_safe_filename(doc_name, ext)
        )
    except (PlaywrightTimeout, Exception):
        print(f"      Could not save from new tab for {doc_name}")
        _cancel_pending_downloads(new_page)
        try:
            return _save_via_request(
                new_page, project_dir, doc_name, fallback_url=fallback_url
            )
        except Exception:
            return None


def _handle_subpage_download(page, btn, project_dir: Path, include_label) -> list[str]:
    """Open a related-documents listing and download matching rows."""
    clear_site_error_toast(page)

    # Use JS click to bypass any modal overlay
    page.evaluate("(el) => el.click()", btn)
    logged_sleep(1)

    return _download_documents_from_current_view(
        page,
        project_dir,
        include_label=include_label,
    )


def _download_documents_from_current_view(page, project_dir: Path, include_label) -> list[str]:
    """Download matching files from the current modal or related-documents page."""

    # Many TOR downloads open a modal popup (bootstrap) with a file list.
    modal_table_ready = False
    try:
        page.wait_for_selector(
            ".modal.show table tbody tr, .modal.fade.show table tbody tr",
            timeout=8_000,
        )
        modal_table_ready = True
    except PlaywrightTimeout:
        modal_table_ready = False

    modal = None
    if modal_table_ready:
        for m in page.query_selector_all(".modal.show, .modal.fade.show"):
            try:
                if not m.is_visible():
                    continue
                ths = m.query_selector_all("th")
                header = " ".join(t.inner_text().strip() for t in ths)
                if "ดาวน์โหลด" in header:
                    modal = m
                    break
            except Exception:
                continue

    download_table = None
    if not modal:
        # Fallback: some TOR pages navigate to a dedicated file-list page.
        try:
            page.wait_for_selector("table", timeout=15_000)
        except PlaywrightTimeout:
            print("      No download table found on sub-page")
            _click_back_or_exit(page)
            return 0

        for t in page.query_selector_all("table"):
            try:
                ths = t.query_selector_all("th")
                header = " ".join(h.inner_text().strip() for h in ths)
                if "ดาวน์โหลด" in header:
                    download_table = t
                    break
            except Exception:
                continue

        if not download_table:
            # As a last resort, proceed with any table rows present.
            try:
                page.wait_for_selector("table tbody tr", timeout=8_000)
            except PlaywrightTimeout:
                print("      No download rows found on sub-page")
                _click_back_or_exit(page)
                return 0

    if modal:
        rows = modal.query_selector_all("table tbody tr")
    elif download_table:
        rows = download_table.query_selector_all("tbody tr")
    else:
        rows = page.query_selector_all("table tbody tr")

    downloaded_labels: list[str] = []
    for row in rows:
        cells = row.query_selector_all("td")
        if len(cells) < 2:
            continue
        cell_texts = [c.inner_text().strip() for c in cells]
        file_label = extract_file_label_from_cell_texts(cell_texts) or "TOR"
        if not include_label(file_label):
            continue

        # Find the download link/button in this row
        last_cell = cells[-1]
        dl_btn = (
            last_cell.query_selector(
                "a[href], a[onclick], button:not([disabled]), [role='button']"
            )
            or last_cell.query_selector("a, button, [role='button']")
            or row.query_selector(
                "a[href], a[onclick], button:not([disabled]), [role='button']"
            )
            or row.query_selector("a, button, [role='button']")
        )

        if dl_btn:
            url_before_click = page.url
            pages_before_click = len(page.context.pages)
            try:

                def _click_and_wait_subpage_download():
                    with page.expect_download(
                        timeout=DOWNLOAD_EVENT_TIMEOUT
                    ) as download_info:
                        page.evaluate("(el) => el.click()", dl_btn)
                    return download_info.value

                download = run_with_toast_recovery(
                    page,
                    _click_and_wait_subpage_download,
                    "Subpage file download",
                    retries=DOWNLOAD_CLICK_RETRIES,
                )
                saved_name = _save_download_to_project(download, project_dir)
                if saved_name:
                    downloaded_labels.append(file_label)
            except PlaywrightTimeout:
                _cancel_pending_downloads(page)
                logged_sleep(0.5)

                # Some rows open a PDF/content page or a new tab instead of a download.
                url_after = page.url
                if url_after != url_before_click:
                    print("      Subpage file opened in content view, saving...")
                    saved_name = _save_from_content_page(page, project_dir, file_label)
                    if saved_name:
                        downloaded_labels.append(file_label)
                    continue

                all_pages = page.context.pages
                if len(all_pages) > pages_before_click:
                    new_page = all_pages[-1]
                    print("      Subpage file opened in new tab, saving...")
                    saved_name = _save_from_new_tab(new_page, project_dir, file_label)
                    try:
                        new_page.close()
                    except Exception:
                        pass
                    if saved_name:
                        downloaded_labels.append(file_label)
                    continue

                print(
                    "      Subpage file download: no event after retries, moving to next file"
                )

    _click_back_or_exit(page)
    return downloaded_labels


def dismiss_modal(page) -> None:
    """Force-dismiss any modal dialog blocking interactions.

    Uses JavaScript evaluate to bypass pointer interception — clicking
    won't work when the modal's own body/footer sits on top of buttons.
    """
    clear_site_error_toast(page)
    page.evaluate("""() => {
        // Remove all visible modals from the DOM entirely
        document.querySelectorAll('.modal.show, .modal.fade.show').forEach(m => {
            m.classList.remove('show');
            m.style.display = 'none';
            m.setAttribute('aria-hidden', 'true');
        });
        // Remove the backdrop overlay too
        document.querySelectorAll('.modal-backdrop').forEach(b => b.remove());
        // Re-enable scrolling on body
        document.body.classList.remove('modal-open');
        document.body.style.removeProperty('overflow');
        document.body.style.removeProperty('padding-right');
    }""")
    logged_sleep(0.5)


def _click_back_or_exit(page) -> None:
    """Click ออก or กลับ button to return to project info page.

    Uses JavaScript click to bypass modal pointer interception —
    the persuadeModalTitle modal can re-appear between dismiss_modal()
    and the actual click, making normal Playwright clicks fail.
    """
    for attempt in range(3):
        dismiss_modal(page)
        clear_site_error_toast(page)
        logged_sleep(0.5)

        # Use JS click which ignores overlapping elements
        clicked = page.evaluate("""() => {
            const labels = ['ออก', 'กลับ', 'ปิด', 'Close'];
            for (const label of labels) {
                const btns = document.querySelectorAll('button');
                for (const btn of btns) {
                    if (btn.textContent.trim().includes(label) && btn.offsetParent !== null) {
                        btn.click();
                        return true;
                    }
                }
            }
            return false;
        }""")
        if clicked:
            logged_sleep(1)
            return

    # Fallback: press Escape
    page.keyboard.press("Escape")
    logged_sleep(1)


def go_back_to_results(page) -> None:
    """Click กลับหน้าหลัก or navigate back to search results.

    Uses JavaScript click to bypass any modal overlays.
    """
    for attempt in range(3):
        dismiss_modal(page)
        clear_site_error_toast(page)
        logged_sleep(0.5)

        clicked = page.evaluate("""() => {
            const labels = ['กลับหน้าหลัก', 'กลับ'];
            // Try buttons first
            for (const label of labels) {
                const btns = document.querySelectorAll('button');
                for (const btn of btns) {
                    if (btn.textContent.trim().includes(label) && btn.offsetParent !== null) {
                        btn.click();
                        return true;
                    }
                }
            }
            // Try links
            for (const label of labels) {
                const links = document.querySelectorAll('a');
                for (const a of links) {
                    if (a.textContent.trim().includes(label) && a.offsetParent !== null) {
                        a.click();
                        return true;
                    }
                }
            }
            return false;
        }""")
        if clicked:
            logged_sleep(2)
            return

    page.go_back()
    logged_sleep(2)


def _process_one_project(
    page,
    row_idx: int,
    preview: str,
    search_name: str,
    processed_projects: dict,
    project_dir_base: Path,
    project_folder_map: dict[str, str] | None = None,
    keyword: str = "",
) -> str:
    """Process a single project: click into it, check eligibility, download docs, update Excel.

    processed_projects is a dict mapping keys → tor_downloaded (bool).
    Keys include project_name, search_name, and project_number.
    Raises Exception on browser crash (caller should handle reconnection).

    Returns:
        "downloaded" — TOR successfully downloaded
        "skipped"    — blacklisted or prelim pricing (no action needed)
        "incomplete" — visited but TOR not available yet
    """
    dismiss_modal(page)

    # Re-query rows (DOM may have changed from previous project)
    rows = get_results_rows(page)
    if row_idx >= len(rows):
        print("      Row index out of range, skipping")
        return "skipped"

    cells = rows[row_idx].query_selector_all("td")
    if len(cells) < 6:
        return "skipped"

    # Click ดูข้อมูล (JS click to bypass any modal overlay)
    view_el = cells[5].query_selector(
        "a[href], a[onclick], button:not([disabled]), [role='button']"
    )
    if not view_el:
        view_el = cells[5].query_selector("a, button, [role='button']")
    if view_el:
        page.evaluate("(el) => el.click()", view_el)
    else:
        page.evaluate("(el) => el.click()", cells[5])
    logged_sleep(3)

    # Check for preliminary pricing → mark done, delete folder, update Excel
    if check_has_preliminary_pricing(page):
        info = extract_project_info(page)
        project_name = info["project_name"] or preview
        info["project_name"] = project_name
        safe_name = sanitize_dirname(project_name)
        project_number = str(info.get("project_number", "") or "").strip()
        if project_folder_map is not None and project_number:
            existing_safe = project_folder_map.get(project_number)
            if existing_safe:
                safe_name = existing_safe
            else:
                project_folder_map[project_number] = safe_name
        project_dir = project_dir_base / safe_name

        print("      Has preliminary pricing — marking complete")
        tracking_status, closed_reason = derive_tracking_status(
            project_name=project_name,
            organization=str(info.get("organization", "") or ""),
            artifact_bucket=ArtifactBucket.NO_ARTIFACT_EVIDENCE,
            prelim_pricing=True,
        )
        info["keyword"] = keyword
        info["search_name"] = search_name
        info["prelim_pricing"] = "Yes"
        info["tor_downloaded"] = "No"
        info["tracking_status"] = tracking_status.value
        info["closed_reason"] = closed_reason.value if closed_reason else ""
        info["artifact_bucket"] = ArtifactBucket.NO_ARTIFACT_EVIDENCE.value
        update_excel(info)

        # Delete downloaded files — bidding is closed, no longer needed
        _safe_remove_dir(project_dir, "Prelim pricing cleanup")

        processed_projects[search_name] = True
        processed_projects[project_name] = True
        if info.get("project_number"):
            processed_projects[info["project_number"]] = True
        return "skipped"

    # Extract project info
    info = extract_project_info(page)
    project_name = info["project_name"] or preview
    info["project_name"] = project_name
    org = info.get("organization", "")

    # Skip projects with blacklisted keywords in name or organization
    if any(sk in project_name or sk in org for sk in SKIP_KEYWORDS_IN_PROJECT):
        print("      Skipping (blacklisted keyword in name/org)")
        processed_projects[search_name] = True
        return "skipped"

    safe_name = sanitize_dirname(project_name)
    project_number = str(info.get("project_number", "") or "").strip()
    if project_folder_map is not None and project_number:
        existing_safe = project_folder_map.get(project_number)
        if existing_safe:
            safe_name = existing_safe
        else:
            project_folder_map[project_number] = safe_name
    project_dir = project_dir_base / safe_name

    print(f"      Project: {project_name[:60]}")
    print(f"      Budget: {info['budget']}")

    # Check if announcement is stale before downloading
    stale = check_announcement_stale(page)

    # Download documents (each doc has its own try/except inside)
    document_summary = download_project_documents(page, project_dir)
    dismiss_modal(page)

    # If announcement is stale and TOR still not available, clean up
    if stale and not document_summary.tor_downloaded:
        print("      Stale announcement — deleting folder, marking complete")
        _safe_remove_dir(project_dir, "Stale announcement cleanup")
        tracking_status, closed_reason = derive_tracking_status(
            project_name=project_name,
            organization=str(org or ""),
            artifact_bucket=document_summary.artifact_bucket,
            stale_without_tor=True,
        )
        info["keyword"] = keyword
        info["search_name"] = search_name
        info["tor_downloaded"] = "No"
        info["prelim_pricing"] = "No"
        info["tracking_status"] = tracking_status.value
        info["closed_reason"] = closed_reason.value if closed_reason else ""
        info["artifact_bucket"] = document_summary.artifact_bucket.value
        update_excel(info)
        processed_projects[search_name] = True
        processed_projects[project_name] = True
        if info.get("project_number"):
            processed_projects[info["project_number"]] = True
        return "skipped"

    # Update Excel with keyword, search_name, and tor_downloaded status
    tracking_status, closed_reason = derive_tracking_status(
        project_name=project_name,
        organization=str(org or ""),
        artifact_bucket=document_summary.artifact_bucket,
    )
    info["keyword"] = keyword
    info["search_name"] = search_name
    info["tor_downloaded"] = "Yes" if document_summary.tor_downloaded else "No"
    info["prelim_pricing"] = "No"
    info["tracking_status"] = tracking_status.value
    info["closed_reason"] = closed_reason.value if closed_reason else ""
    info["artifact_bucket"] = document_summary.artifact_bucket.value
    update_excel(info)
    if project_dir.exists():
        write_project_manifest(
            project_dir=project_dir,
            project_info=info,
            tracking_status=tracking_status,
            closed_reason=closed_reason,
            artifact_bucket=document_summary.artifact_bucket,
        )
    processed_projects[search_name] = document_summary.tor_downloaded
    processed_projects[project_name] = document_summary.tor_downloaded
    if info.get("project_number"):
        processed_projects[info["project_number"]] = document_summary.tor_downloaded
    return "downloaded" if document_summary.tor_downloaded else "incomplete"


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------


def parse_cli_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Parse command-line arguments for the crawler.

    Returns a namespace with `.profile` set to "tor" (default), "toe", or "lue".
    Accepts an explicit argv list so the function is unit-testable without
    touching sys.argv.
    """
    parser = argparse.ArgumentParser(
        description="e-GP Thailand Government Procurement Crawler",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Profiles:\n"
            "  tor  IT/data/consulting projects  →  OneDrive/Download/TOR  (default)\n"
            "  toe  AV/display equipment         →  OneDrive/Download/TOE\n"
            "  lue  PR/ads/training/web/online   →  OneDrive/Download/LUE\n"
        ),
    )
    parser.add_argument(
        "--profile",
        choices=("tor", "toe", "lue"),
        default="tor",
        help="Keyword/output profile to run (default: tor)",
    )
    return parser.parse_args(argv)


def main(profile: str = "tor") -> None:
    """Run the full crawling workflow across all keywords."""
    apply_profile_defaults(profile)
    dotenv_path = load_dotenv_from_default_locations(override=False)
    apply_env_config_overrides()
    configure_runtime_paths()
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    # Load previously downloaded projects: dict of project_name → tor_downloaded (bool)
    existing_projects = load_existing_projects()
    project_folder_map = load_existing_project_folder_map()
    total_project_rows, complete_rows, incomplete_rows = (
        load_existing_project_row_stats()
    )
    processed_projects: dict[str, bool] = dict(existing_projects)
    total_processed = 0
    total_tor_downloaded = 0

    print("=" * 60)
    print("e-GP Procurement Crawler")
    print(f"Download directory: {DOWNLOAD_DIR}")
    print(f"Browser profile: {BROWSER_PROFILE_DIR}")
    print(f"Excel file: {EXCEL_PATH}")
    if dotenv_path:
        print(f"Env file loaded: {dotenv_path}")
    print(f"Keywords to search: {len(KEYWORDS)}")
    if total_project_rows:
        print(
            f"Existing projects: {total_project_rows} "
            f"({complete_rows} complete, {incomplete_rows} need TOR check)"
        )
    print("=" * 60)

    print("\nLaunching Chrome browser...", flush=True)
    chrome_proc = launch_real_chrome()

    pw = sync_playwright().start()
    browser, page = connect_playwright_to_chrome(pw)
    print("Connected to Chrome via CDP.\n", flush=True)

    try:
        # Step 1: Go to main page first to warm up Cloudflare cookies
        print("Step 1: Opening main e-GP page...")
        page.goto(MAIN_PAGE_URL, wait_until="domcontentloaded", timeout=NAV_TIMEOUT)
        logged_sleep(3)
        wait_for_cloudflare(page)

        # Step 2: Navigate to the search page
        print("\nStep 2: Opening search page...")
        page.goto(SEARCH_URL, wait_until="domcontentloaded", timeout=NAV_TIMEOUT)
        logged_sleep(5)
        wait_for_cloudflare(page)
        print("Ready to search.\n")

        kw_idx = 0
        resume_state: KeywordResumeState | None = None
        while kw_idx < len(KEYWORDS):
            keyword = KEYWORDS[kw_idx]
            print(f"\n[{kw_idx + 1}/{len(KEYWORDS)}] Keyword: {keyword}")
            print("-" * 40)
            page_num = 1

            try:
                if resume_state is not None and resume_state.keyword_index == kw_idx:
                    restore_results_page(page, keyword, resume_state.page_num)
                    page_num = resume_state.page_num
                else:
                    if kw_idx > 0:
                        clear_search(page)

                    search_keyword(page, keyword)

                    # Check if search returned any results at all (avoid false positives from body text)
                    if is_no_results_page(page):
                        print("    No results for this keyword")
                        log_results_debug_snapshot(page, keyword, "no-results")
                        kw_idx += 1
                        resume_state = None
                        continue

                # Process page-by-page to avoid stale DOM references
                while page_num <= MAX_PAGES_PER_KEYWORD:
                    print(f"    Scanning results page {page_num}...")
                    rows = get_results_rows(page)
                    eligible_indices = []

                    for i, row in enumerate(rows):
                        cells = row.query_selector_all("td")
                        if len(cells) < 6:
                            continue
                        status_text = cells[4].inner_text().strip()
                        if status_matches_target(status_text):
                            project_text = cells[2].inner_text().strip()
                            project_preview = project_text[:80]

                            # Skip projects with blacklisted keywords
                            if any(
                                sk in project_text for sk in SKIP_KEYWORDS_IN_PROJECT
                            ):
                                continue

                            # Check dedup: skip only if TOR already downloaded
                            # Exact lookup by search-table name (O(1) dict lookup)
                            if processed_projects.get(project_text, False):
                                continue  # TOR downloaded — fully done
                            eligible_indices.append((i, project_preview, project_text))

                    print(
                        f"    Found {len(eligible_indices)} new eligible projects on this page"
                    )

                    # Process each eligible project
                    for idx, (row_idx, preview, search_name) in enumerate(
                        eligible_indices
                    ):
                        print(
                            f"\n      [{idx + 1}/{len(eligible_indices)}] {preview}..."
                        )

                        try:
                            result = _process_one_project(
                                page,
                                row_idx,
                                preview,
                                search_name,
                                processed_projects,
                                project_dir_base=DOWNLOAD_DIR,
                                project_folder_map=project_folder_map,
                                keyword=keyword,
                            )
                            total_processed += 1
                            if result == "downloaded":
                                total_tor_downloaded += 1
                            print(
                                f"      Done ({result}). "
                                f"Processed: {total_processed}, TOR downloaded: {total_tor_downloaded}"
                            )
                        except Exception as proj_err:
                            err_msg = str(proj_err)
                            if "has been closed" in err_msg:
                                raise  # Bubble up browser crash to keyword handler
                            print(f"      ERROR on project, skipping: {err_msg}")
                            # Don't mark as TOR-complete on error — allow revisit
                            if search_name not in processed_projects:
                                processed_projects[search_name] = False

                        # Always try to go back to results after each project
                        try:
                            go_back_to_results(page)
                            logged_sleep(2)
                        except Exception:
                            pass

                    # Check if there are page numbers (1, 2, 3...) — no numbers means
                    # no more pages. We rely on ถัดไป being present+enabled rather than
                    # specific tag names for numeric page links.
                    dismiss_modal(page)
                    previous_marker = get_results_page_marker(page)
                    next_btn = page.query_selector(
                        "a:has-text('ถัดไป'), button:has-text('ถัดไป')"
                    )
                    if not (next_btn and next_btn.is_visible()):
                        break

                    try:
                        state = next_btn.evaluate(
                            """el => {
                                const li = el.closest('li');
                                const src = li || el;
                                return {
                                    ariaDisabled: el.getAttribute('aria-disabled') ||
                                                 (src && src.getAttribute ? src.getAttribute('aria-disabled') : null),
                                    disabled: ('disabled' in el) ? el.disabled : null,
                                    className: (src && src.className) ? String(src.className) : '',
                                };
                            }"""
                        )
                    except Exception:
                        state = None

                    if state and pagination_button_is_disabled(
                        state.get("ariaDisabled"),
                        state.get("disabled"),
                        state.get("className"),
                    ):
                        break

                    try:
                        next_btn.scroll_into_view_if_needed()
                    except Exception:
                        pass

                    try:
                        page.evaluate("(el) => el.click()", next_btn)
                    except Exception:
                        try:
                            next_btn.click(timeout=10_000)
                        except Exception:
                            break

                    logged_sleep(3)
                    if not wait_for_results_page_change(page, previous_marker):
                        break
                    # Check if we landed on an empty page
                    if is_no_results_page(page):
                        break
                    page_num += 1

                resume_state = None
                kw_idx += 1

            except Exception as e:
                error_msg = str(e)
                print(f"  ERROR processing keyword '{keyword}': {error_msg}")

                if "has been closed" in error_msg:
                    resume_state = KeywordResumeState(
                        keyword_index=kw_idx,
                        keyword=keyword,
                        page_num=page_num,
                    )
                    # Chrome crashed — restart browser and reconnect
                    print("  Chrome crashed. Restarting browser...")
                    try:
                        browser.close()
                    except Exception:
                        pass
                    try:
                        pw.stop()
                    except Exception:
                        pass
                    try:
                        chrome_proc.kill()
                        chrome_proc.wait(timeout=5)
                    except Exception:
                        pass

                    logged_sleep(3)
                    chrome_proc = launch_real_chrome()
                    pw = sync_playwright().start()
                    browser, page = connect_playwright_to_chrome(pw)
                    print("  Reconnected to Chrome.")

                    # Navigate to search page
                    page.goto(
                        MAIN_PAGE_URL,
                        wait_until="domcontentloaded",
                        timeout=NAV_TIMEOUT,
                    )
                    logged_sleep(3)
                    wait_for_cloudflare(page)
                    page.goto(
                        SEARCH_URL, wait_until="domcontentloaded", timeout=NAV_TIMEOUT
                    )
                    logged_sleep(5)
                    wait_for_cloudflare(page)
                    continue
                else:
                    resume_state = None
                    kw_idx += 1
                    # Non-crash error — try to navigate back to search
                    try:
                        page.goto(
                            SEARCH_URL,
                            wait_until="domcontentloaded",
                            timeout=NAV_TIMEOUT,
                        )
                        logged_sleep(3)
                        wait_for_cloudflare(page)
                    except Exception:
                        pass

    finally:
        safe_shutdown(
            browser=browser, pw=pw, chrome_proc=chrome_proc, ignore_sigint=True
        )
        sync_excel_back_to_onedrive()

    print("\n" + "=" * 60)
    print("Crawling complete.")
    print(f"  Projects processed: {total_processed}")
    print(f"  TOR downloaded:     {total_tor_downloaded}")
    print(f"  Excel file: {EXCEL_PATH}")
    print("=" * 60)


if __name__ == "__main__":
    _args = parse_cli_args()
    main(profile=_args.profile)
