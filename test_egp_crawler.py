"""Unit tests for egp_crawler.py — tests pure functions without browser."""

from datetime import datetime, date
import json
from pathlib import Path
import socket

from egp_shared_types.enums import ArtifactBucket, ClosedReason, ProjectState
from openpyxl import Workbook, load_workbook

import os

from egp_crawler import (
    DOCS_TO_DOWNLOAD,
    DOWNLOAD_TIMEOUT,
    EXCEL_HEADERS,
    KEYWORDS,
    KEYWORDS_DEFAULT,
    KEYWORDS_TOE_DEFAULT,
    KEYWORDS_LUE_DEFAULT,
    PROFILE_DEFAULTS,
    PlaywrightTimeout,
    SKIP_KEYWORDS_IN_PROJECT,
    SUBPAGE_DOWNLOAD_TIMEOUT,
    apply_profile_defaults,
    parse_cli_args,
    build_safe_filename,
    build_results_debug_snapshot,
    download_project_documents,
    env_get_float,
    env_get_int,
    env_get_path,
    env_get_str,
    extract_document_url_from_viewer_url,
    extract_file_label_from_cell_texts,
    extract_url_from_onclick,
    is_allowed_download_url,
    is_show_htmlfile_url,
    has_site_error_toast_text,
    is_tor_doc_label,
    is_tor_file,
    looks_like_html_bytes,
    load_existing_project_row_stats,
    load_existing_projects,
    filename_from_content_disposition,
    get_results_page_marker,
    _goto_with_retries,
    keywords_from_env,
    load_dotenv_file,
    pagination_button_is_disabled,
    parse_buddhist_date,
    parse_keywords,
    results_page_marker_changed,
    restore_results_page,
    safe_shutdown,
    sanitize_dirname,
    sanitize_filename_preserve_suffix,
    SEARCH_URL,
    sniff_extension_from_bytes,
    status_matches_target,
    _download_one_document,
    _save_from_new_tab,
    wait_for_cloudflare,
    update_excel,
    wait_for_local_tcp_listen,
    wait_for_results_ready,
    write_project_manifest,
    search_keyword,
)


# ---------------------------------------------------------------------------
# sanitize_dirname
# ---------------------------------------------------------------------------


class TestSanitizeDirname:
    def test_removes_illegal_chars(self):
        assert sanitize_dirname('a/b\\c:d*e?f"g<h>i|j') == "abcdefghij"

    def test_preserves_thai_text(self):
        thai = "โครงการวิเคราะห์ข้อมูล"
        assert sanitize_dirname(thai) == thai

    def test_truncates_long_names(self):
        long_name = "ก" * 200
        result = sanitize_dirname(long_name)
        assert len(result) == 100

    def test_strips_whitespace(self):
        assert sanitize_dirname("  hello world  ") == "hello world"

    def test_collapses_multiple_spaces(self):
        assert sanitize_dirname("hello   world   test") == "hello world test"

    def test_handles_newlines(self):
        assert sanitize_dirname("line1\nline2\rline3") == "line1 line2 line3"

    def test_empty_string(self):
        assert sanitize_dirname("") == ""

    def test_mixed_thai_english_with_illegal(self):
        name = 'โครงการ "test" <project>'
        assert sanitize_dirname(name) == "โครงการ test project"


class TestSanitizeFilenamePreserveSuffix:
    def test_truncates_but_keeps_pdf_suffix(self):
        name = ("a" * 150) + ".pdf"
        result = sanitize_filename_preserve_suffix(name, max_len=100)
        assert result.endswith(".pdf")
        assert len(result) == 100

    def test_truncates_but_keeps_multi_suffix(self):
        name = ("a" * 150) + ".tar.gz"
        result = sanitize_filename_preserve_suffix(name, max_len=100)
        assert result.endswith(".tar.gz")
        assert len(result) == 100

    def test_truncates_no_suffix(self):
        name = "a" * 200
        result = sanitize_filename_preserve_suffix(name, max_len=100)
        assert len(result) == 100


class TestBuildSafeFilename:
    def test_truncates_stem_to_fit_extension(self):
        result = build_safe_filename("a" * 200, ".pdf", max_len=100)
        assert result.endswith(".pdf")
        assert len(result) == 100

    def test_adds_dot_when_missing(self):
        assert build_safe_filename("name", "pdf", max_len=100).endswith(".pdf")

    def test_uses_default_stem_when_empty(self):
        assert build_safe_filename("", ".pdf", max_len=20).endswith(".pdf")

    def test_only_illegal_chars(self):
        assert sanitize_dirname('/:*?"<>|\\') == ""


# ---------------------------------------------------------------------------
# update_excel
# ---------------------------------------------------------------------------


class TestUpdateExcel:
    def test_creates_new_file_with_headers(self, tmp_path):
        excel_path = tmp_path / "test.xlsx"
        info = {
            "download_date": "2025-01-15",
            "project_name": "Test Project",
            "organization": "Test Org",
            "project_number": "12345",
            "budget": "1,000,000.00",
            "proposal_submission_date": "20/01/2568",
        }
        update_excel(info, excel_path)

        wb = load_workbook(excel_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert headers == EXCEL_HEADERS

    def test_appends_to_existing(self, tmp_path):
        excel_path = tmp_path / "test.xlsx"

        info1 = {
            "download_date": "2025-01-15",
            "project_name": "Project A",
            "organization": "Org A",
            "project_number": "111",
            "budget": "100",
            "proposal_submission_date": "01/01/2568",
        }
        info2 = {
            "download_date": "2025-01-16",
            "project_name": "Project B",
            "organization": "Org B",
            "project_number": "222",
            "budget": "200",
            "proposal_submission_date": "02/01/2568",
        }

        update_excel(info1, excel_path)
        update_excel(info2, excel_path)

        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws.max_row == 3  # header + 2 data rows
        assert ws.cell(2, 2).value == "Project A"
        assert ws.cell(3, 2).value == "Project B"

    def test_correct_column_count(self, tmp_path):
        excel_path = tmp_path / "test.xlsx"
        info = {
            "download_date": "2025-01-15",
            "project_name": "Test",
            "organization": "Org",
            "project_number": "123",
            "budget": "500",
            "proposal_submission_date": "15/01/2568",
            "keyword": "ระบบสารสนเทศ",
            "tor_downloaded": "Yes",
            "prelim_pricing": "No",
            "search_name": "Test from table",
        }
        update_excel(info, excel_path)

        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws.max_column == len(EXCEL_HEADERS)

    def test_handles_missing_fields(self, tmp_path):
        excel_path = tmp_path / "test.xlsx"
        info = {"project_name": "Partial"}
        update_excel(info, excel_path)

        wb = load_workbook(excel_path)
        ws = wb.active
        # Should still create a row with defaults
        assert ws.cell(2, 2).value == "Partial"
        assert ws.cell(2, 3).value in ("", None)  # openpyxl stores "" as None

    def test_download_date_defaults_to_today(self, tmp_path):
        excel_path = tmp_path / "test.xlsx"
        info = {"project_name": "Auto Date"}
        update_excel(info, excel_path)

        wb = load_workbook(excel_path)
        ws = wb.active
        today = datetime.now().strftime("%Y-%m-%d")
        assert ws.cell(2, 1).value == today

    def test_preserves_existing_data_on_append(self, tmp_path):
        excel_path = tmp_path / "test.xlsx"

        for i in range(5):
            update_excel({"project_name": f"Project {i}"}, excel_path)

        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws.max_row == 6  # header + 5 rows
        for i in range(5):
            assert ws.cell(i + 2, 2).value == f"Project {i}"

    def test_creates_parent_directories(self, tmp_path):
        excel_path = tmp_path / "deep" / "nested" / "dir" / "test.xlsx"
        update_excel({"project_name": "Deep"}, excel_path)
        assert excel_path.exists()

    def test_backfills_legacy_9_column_headers(self, tmp_path):
        """Existing old sheets should be upgraded to 10-column headers."""
        excel_path = tmp_path / "legacy.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Projects"
        ws.append(EXCEL_HEADERS[:9])  # simulate old file without search_name
        ws.append(["2025-01-01", "Legacy Project", "", "", "", "", "", "No", "No"])
        wb.save(excel_path)

        update_excel(
            {
                "project_name": "Legacy Project",
                "search_name": "Legacy from table",
                "tor_downloaded": "Yes",
            },
            excel_path,
        )

        wb2 = load_workbook(excel_path)
        ws2 = wb2.active
        headers = [ws2.cell(1, i).value for i in range(1, len(EXCEL_HEADERS) + 1)]
        assert headers == EXCEL_HEADERS
        assert ws2.cell(2, 10).value == "Legacy from table"

    def test_updates_existing_row_by_project_number(self, tmp_path):
        excel_path = tmp_path / "by_number.xlsx"
        update_excel(
            {
                "project_name": "Old Name",
                "project_number": "6901",
                "tor_downloaded": "No",
            },
            excel_path,
        )
        # Same project_number, different name should update in place (not append)
        update_excel(
            {
                "project_name": "New Name (should not duplicate)",
                "project_number": "6901",
                "tor_downloaded": "Yes",
                "search_name": "From results table",
            },
            excel_path,
        )

        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws.max_row == 2  # header + 1 row
        assert ws.cell(2, 4).value == "6901"
        assert ws.cell(2, 8).value == "Yes"
        assert ws.cell(2, 10).value == "From results table"

    def test_reannouncement_with_same_name_but_new_project_number_appends(self, tmp_path):
        excel_path = tmp_path / "reannounce.xlsx"
        update_excel(
            {
                "project_name": "โครงการระบบข้อมูลกลาง",
                "project_number": "69010000001",
                "search_name": "โครงการระบบข้อมูลกลาง (เลขที่โครงการ : 69010000001)",
                "tor_downloaded": "Yes",
                "tracking_status": ProjectState.TOR_DOWNLOADED.value,
            },
            excel_path,
        )
        update_excel(
            {
                "project_name": "โครงการระบบข้อมูลกลาง",
                "project_number": "69020000002",
                "search_name": "โครงการระบบข้อมูลกลาง (เลขที่โครงการ : 69020000002)",
                "tor_downloaded": "No",
                "tracking_status": ProjectState.OPEN_INVITATION.value,
            },
            excel_path,
        )

        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws.max_row == 3
        assert ws.cell(2, 4).value == "69010000001"
        assert ws.cell(3, 4).value == "69020000002"

    def test_backfills_numberless_row_when_same_project_revisited(self, tmp_path):
        excel_path = tmp_path / "backfill.xlsx"
        update_excel(
            {
                "project_name": "โครงการระบบข้อมูลกลาง",
                "search_name": "โครงการระบบข้อมูลกลาง",
                "tor_downloaded": "No",
            },
            excel_path,
        )
        update_excel(
            {
                "project_name": "โครงการระบบข้อมูลกลาง",
                "project_number": "69030000003",
                "search_name": "โครงการระบบข้อมูลกลาง",
                "tor_downloaded": "Yes",
                "tracking_status": ProjectState.TOR_DOWNLOADED.value,
            },
            excel_path,
        )

        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws.max_row == 2
        assert ws.cell(2, 4).value == "69030000003"
        assert ws.cell(2, 8).value == "Yes"


# ---------------------------------------------------------------------------
# load_existing_projects (deduplication)
# ---------------------------------------------------------------------------


class TestLoadExistingProjects:
    def test_returns_empty_dict_when_no_file(self, tmp_path):
        result = load_existing_projects(tmp_path / "nonexistent.xlsx")
        assert result == {}

    def test_loads_project_names_with_tor_status(self, tmp_path):
        excel_path = tmp_path / "test.xlsx"
        update_excel({"project_name": "Project A", "tor_downloaded": "Yes"}, excel_path)
        update_excel({"project_name": "Project B", "tor_downloaded": "No"}, excel_path)
        update_excel(
            {"project_name": "โครงการทดสอบ", "tor_downloaded": "Yes"}, excel_path
        )

        result = load_existing_projects(excel_path)
        assert result["Project A"] is True
        assert result["Project B"] is False
        assert result["โครงการทดสอบ"] is True


# ---------------------------------------------------------------------------
# extract_file_label_from_cell_texts
# ---------------------------------------------------------------------------


class TestExtractFileLabelFromCellTexts:
    def test_picks_extension_token(self):
        texts = ["1", "ฉบับแรก 68109150577_01122568.zip", "01/12/2568", "04/12/2568", ""]
        assert (
            extract_file_label_from_cell_texts(texts)
            == "ฉบับแรก 68109150577_01122568.zip"
        )

    def test_falls_back_to_second_column_when_first_is_index(self):
        assert (
            extract_file_label_from_cell_texts(["2", "68109150577_16012569_1.zip"])
            == "68109150577_16012569_1.zip"
        )

    def test_falls_back_to_first_when_not_index(self):
        assert (
            extract_file_label_from_cell_texts(["pricebuild_123.zip", ""])
            == "pricebuild_123.zip"
        )

    def test_empty(self):
        assert extract_file_label_from_cell_texts([]) == ""

    def test_skips_empty_project_names(self, tmp_path):
        excel_path = tmp_path / "test.xlsx"
        update_excel(
            {"project_name": "Real Project", "tor_downloaded": "No"}, excel_path
        )
        update_excel({"project_name": ""}, excel_path)

        result = load_existing_projects(excel_path)
        assert "Real Project" in result
        assert "" not in result

    def test_incomplete_projects_not_skipped(self, tmp_path):
        """Projects without TOR downloaded should be revisited."""
        excel_path = tmp_path / "test.xlsx"
        update_excel({"project_name": "Incomplete", "tor_downloaded": "No"}, excel_path)

        existing = load_existing_projects(excel_path)
        assert "Incomplete" in existing
        assert existing["Incomplete"] is False  # Not complete — should revisit

    def test_complete_projects_skipped(self, tmp_path):
        """Projects with TOR downloaded should be fully skipped."""
        excel_path = tmp_path / "test.xlsx"
        update_excel({"project_name": "Complete", "tor_downloaded": "Yes"}, excel_path)

        existing = load_existing_projects(excel_path)
        assert existing["Complete"] is True  # Complete — skip

    def test_keyword_column_stored(self, tmp_path):
        excel_path = tmp_path / "test.xlsx"
        update_excel(
            {
                "project_name": "With KW",
                "keyword": "ระบบสารสนเทศ",
                "tor_downloaded": "No",
            },
            excel_path,
        )

        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws.cell(2, 7).value == "ระบบสารสนเทศ"

    def test_update_existing_project_tor_status(self, tmp_path):
        """When revisiting, tor_downloaded should be updated in-place."""
        excel_path = tmp_path / "test.xlsx"
        update_excel({"project_name": "Revisit Me", "tor_downloaded": "No"}, excel_path)

        # Simulate re-download with TOR now available
        update_excel(
            {"project_name": "Revisit Me", "tor_downloaded": "Yes"}, excel_path
        )

        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws.max_row == 2  # Still only 1 data row (updated in place)
        assert ws.cell(2, 8).value == "Yes"

    def test_prelim_pricing_skips_project(self, tmp_path):
        """Projects with prelim_pricing=Yes should be fully skipped."""
        excel_path = tmp_path / "test.xlsx"
        update_excel(
            {
                "project_name": "Prelim Done",
                "tor_downloaded": "Yes",
                "prelim_pricing": "Yes",
            },
            excel_path,
        )

        existing = load_existing_projects(excel_path)
        assert existing["Prelim Done"] is True

    def test_prelim_pricing_updated_in_place(self, tmp_path):
        """When prelim pricing found on revisit, update in place."""
        excel_path = tmp_path / "test.xlsx"
        update_excel(
            {
                "project_name": "Was Active",
                "tor_downloaded": "No",
                "prelim_pricing": "No",
            },
            excel_path,
        )

        # Simulate revisit finding prelim pricing
        update_excel(
            {
                "project_name": "Was Active",
                "tor_downloaded": "Yes",
                "prelim_pricing": "Yes",
            },
            excel_path,
        )

        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws.max_row == 2  # No duplicate
        assert ws.cell(2, 9).value == "Yes"  # prelim_pricing updated

    def test_row_stats_count_unique_projects_not_keys(self, tmp_path):
        excel_path = tmp_path / "test.xlsx"
        update_excel(
            {
                "project_name": "A",
                "project_number": "100",
                "search_name": "A table",
                "tor_downloaded": "Yes",
            },
            excel_path,
        )
        update_excel(
            {
                "project_name": "B",
                "project_number": "200",
                "search_name": "B table",
                "tor_downloaded": "No",
            },
            excel_path,
        )
        total, complete, incomplete = load_existing_project_row_stats(excel_path)
        assert total == 2
        assert complete == 1
        assert incomplete == 1


# ---------------------------------------------------------------------------
# Constants validation
# ---------------------------------------------------------------------------


class TestConstants:
    def test_keywords_list_has_twelve_entries(self):
        assert len(KEYWORDS) == 12

    def test_keywords_are_nonempty_strings(self):
        for kw in KEYWORDS:
            assert isinstance(kw, str)
            assert len(kw) > 0

    def test_docs_to_download_has_four_entries(self):
        assert len(DOCS_TO_DOWNLOAD) == 4

    def test_excel_headers_has_ten_entries(self):
        assert len(EXCEL_HEADERS) == 13

    def test_excel_headers_include_keyword_tor_prelim_search_name(self):
        assert "keyword" in EXCEL_HEADERS
        assert "tor_downloaded" in EXCEL_HEADERS
        assert "prelim_pricing" in EXCEL_HEADERS
        assert "search_name" in EXCEL_HEADERS
        assert "tracking_status" in EXCEL_HEADERS
        assert "closed_reason" in EXCEL_HEADERS
        assert "artifact_bucket" in EXCEL_HEADERS

    def test_keywords_contain_expected_terms(self):
        expected = {"วิเคราะห์ข้อมูล", "ระบบสารสนเทศ", "ที่ปรึกษา", "เทคโนโลยีสารสนเทศ"}
        assert expected.issubset(set(KEYWORDS))

    def test_skip_keywords_include_maintenance(self):
        assert "บำรุงรักษา" in SKIP_KEYWORDS_IN_PROJECT


# ---------------------------------------------------------------------------
# search_name column (multi-key dedup)
# ---------------------------------------------------------------------------


class TestSearchNameDedup:
    def test_search_name_stored_in_excel(self, tmp_path):
        """search_name should be written to column J (index 10)."""
        excel_path = tmp_path / "test.xlsx"
        update_excel(
            {
                "project_name": "Detail Page Name",
                "search_name": "Table Name",
                "tor_downloaded": "No",
            },
            excel_path,
        )

        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws.cell(2, 10).value == "Table Name"

    def test_load_indexes_by_search_name(self, tmp_path):
        """load_existing_projects should return entries keyed by search_name too."""
        excel_path = tmp_path / "test.xlsx"
        update_excel(
            {
                "project_name": "Detail Name",
                "search_name": "Table Name",
                "tor_downloaded": "Yes",
            },
            excel_path,
        )

        result = load_existing_projects(excel_path)
        assert "Table Name" in result
        assert result["Table Name"] is True

    def test_load_indexes_by_project_number(self, tmp_path):
        """load_existing_projects should return entries keyed by project_number too."""
        excel_path = tmp_path / "test.xlsx"
        update_excel(
            {
                "project_name": "Some Project",
                "project_number": "67129000001",
                "tor_downloaded": "Yes",
            },
            excel_path,
        )

        result = load_existing_projects(excel_path)
        assert "67129000001" in result
        assert result["67129000001"] is True

    def test_all_three_keys_share_same_status(self, tmp_path):
        """All keys for the same project should report the same completion status."""
        excel_path = tmp_path / "test.xlsx"
        update_excel(
            {
                "project_name": "Detail Name",
                "search_name": "Table Name",
                "project_number": "99001",
                "tor_downloaded": "No",
            },
            excel_path,
        )

        result = load_existing_projects(excel_path)
        assert result["Table Name"] is False
        assert result["99001"] is False

    def test_missing_search_name_only_indexes_by_name(self, tmp_path):
        """If search_name is empty, don't add empty string as a key."""
        excel_path = tmp_path / "test.xlsx"
        update_excel(
            {
                "project_name": "Only Name",
                "tor_downloaded": "Yes",
            },
            excel_path,
        )

        result = load_existing_projects(excel_path)
        assert "Only Name" in result
        assert "" not in result

    def test_tracking_status_controls_terminal_skip(self, tmp_path):
        excel_path = tmp_path / "status.xlsx"
        update_excel(
            {
                "project_name": "Consulting timeout",
                "search_name": "Consulting timeout",
                "tor_downloaded": "No",
                "tracking_status": ProjectState.CLOSED_TIMEOUT_CONSULTING.value,
                "closed_reason": ClosedReason.CONSULTING_TIMEOUT_30D.value,
            },
            excel_path,
        )

        result = load_existing_projects(excel_path)
        assert result["Consulting timeout"] is True

    def test_missing_project_number_only_indexes_by_name(self, tmp_path):
        """If project_number is empty, don't add empty string as a key."""
        excel_path = tmp_path / "test.xlsx"
        update_excel(
            {
                "project_name": "Only Name",
                "project_number": "",
                "tor_downloaded": "Yes",
            },
            excel_path,
        )

        result = load_existing_projects(excel_path)
        assert "Only Name" in result
        # Empty project_number should NOT be a key
        keys_list = list(result.keys())
        assert all(k != "" for k in keys_list)

    def test_search_name_updated_on_revisit(self, tmp_path):
        """When updating existing project, search_name should be updated."""
        excel_path = tmp_path / "test.xlsx"
        update_excel(
            {
                "project_name": "My Project",
                "tor_downloaded": "No",
            },
            excel_path,
        )

        # Revisit with search_name now available
        update_excel(
            {
                "project_name": "My Project",
                "search_name": "Table Version",
                "tor_downloaded": "Yes",
            },
            excel_path,
        )

        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws.max_row == 2  # No duplicate
        assert ws.cell(2, 10).value == "Table Version"

    def test_no_false_positive_on_similar_names(self, tmp_path):
        """Two projects with similar prefixes should NOT collide."""
        excel_path = tmp_path / "test.xlsx"
        prefix = "ประกวดราคาจ้างโครงการพัฒนาระบบสารสนเทศเพื่อการบริหาร"
        update_excel(
            {
                "project_name": prefix + "งานบุคคล",
                "search_name": prefix + "งานบุคคล (table)",
                "tor_downloaded": "Yes",
            },
            excel_path,
        )

        result = load_existing_projects(excel_path)
        # The OTHER project with a similar prefix should NOT match
        assert (prefix + "จัดการงบประมาณ") not in result
        assert (prefix + "จัดการงบประมาณ (table)") not in result


# ---------------------------------------------------------------------------
# parse_buddhist_date
# ---------------------------------------------------------------------------


class TestParseBuddhistDate:
    def test_valid_date(self):
        """30/12/2568 (Buddhist) = 30 Dec 2025 (Gregorian)."""
        result = parse_buddhist_date("30/12/2568")
        assert result == date(2025, 12, 30)

    def test_another_valid_date(self):
        """15/01/2569 (Buddhist) = 15 Jan 2026 (Gregorian)."""
        result = parse_buddhist_date("15/01/2569")
        assert result == date(2026, 1, 15)

    def test_returns_none_on_empty(self):
        assert parse_buddhist_date("") is None

    def test_returns_none_on_garbage(self):
        assert parse_buddhist_date("not-a-date") is None

    def test_returns_none_on_partial(self):
        assert parse_buddhist_date("30/12") is None

    def test_strips_whitespace(self):
        result = parse_buddhist_date("  30/12/2568  ")
        assert result == date(2025, 12, 30)


# ---------------------------------------------------------------------------
# is_tor_file (distinguishes TOR from pricebuild)
# ---------------------------------------------------------------------------


class TestIsTorFile:
    def test_pricebuild_zip_is_not_tor(self):
        assert is_tor_file("pricebuild_2600700000_69029018854.zip") is False

    def test_pricebuild_with_different_numbers_is_not_tor(self):
        assert is_tor_file("pricebuild_141040070010000035_69019456654.zip") is False

    def test_tor_pdf_is_tor(self):
        assert is_tor_file("TOR_chapter1.pdf") is True

    def test_regular_zip_is_tor(self):
        assert is_tor_file("document_69029018854.zip") is True

    def test_docx_is_tor(self):
        assert is_tor_file("ขอบเขตของงาน.docx") is True

    def test_pB0_pdf_is_not_tor(self):
        """pB0.pdf is a price build-up document, not TOR."""
        assert is_tor_file("pB0.pdf") is False

    def test_pB_with_number_is_not_tor(self):
        assert is_tor_file("pB1.pdf") is False
        assert is_tor_file("pB12.pdf") is False
        assert is_tor_file("B1.pdf") is False
        assert is_tor_file("B12.PDF") is False

    def test_empty_filename_is_not_tor(self):
        assert is_tor_file("") is False


class TestIsTorDocLabel:
    def test_matches_consulting_tor_label(self):
        assert is_tor_doc_label("ร่างขอบเขตของงาน (TOR)") is True

    def test_matches_consulting_document_label(self):
        assert is_tor_doc_label("ร่างเอกสารจ้างที่ปรึกษา") is True

    def test_does_not_match_price_announcement(self):
        assert is_tor_doc_label("ประกาศราคากลาง") is False


# ---------------------------------------------------------------------------
# Viewer URL helpers
# ---------------------------------------------------------------------------


class TestExtractDocumentUrlFromViewerUrl:
    def test_extracts_file_param(self):
        url = (
            "chrome-extension://mhjfbmdgcfjbbpaeojofohoefgiehjai/index.html"
            "?file=https%3A%2F%2Fexample.com%2Fdoc.pdf"
        )
        assert (
            extract_document_url_from_viewer_url(url) == "https://example.com/doc.pdf"
        )

    def test_extracts_url_param(self):
        url = (
            "chrome-extension://ext/index.html?url=https%3A%2F%2Fexample.com%2Ftor.zip"
        )
        assert (
            extract_document_url_from_viewer_url(url) == "https://example.com/tor.zip"
        )

    def test_extracts_any_param_value_that_is_http(self):
        url = "chrome-extension://ext/index.html?foo=bar&x=https%3A%2F%2Fexample.com%2Fa.pdf"
        assert extract_document_url_from_viewer_url(url) == "https://example.com/a.pdf"

    def test_returns_none_when_no_param(self):
        assert extract_document_url_from_viewer_url("https://example.com/view") is None


class TestFilenameFromContentDisposition:
    def test_parses_simple_filename(self):
        header = 'attachment; filename="TOR.pdf"'
        assert filename_from_content_disposition(header) == "TOR.pdf"

    def test_parses_rfc5987_filename_star(self):
        header = "attachment; filename*=UTF-8''TOR%20Thai.pdf"
        assert filename_from_content_disposition(header) == "TOR Thai.pdf"


class TestIsAllowedDownloadUrl:
    def test_allows_gprocurement_hosts(self):
        assert (
            is_allowed_download_url("https://process5.gprocurement.go.th/a.pdf") is True
        )
        assert (
            is_allowed_download_url("https://www.gprocurement.go.th/egp/file.zip")
            is True
        )

    def test_blocks_non_gprocurement_hosts(self):
        assert is_allowed_download_url("https://example.com/a.pdf") is False

    def test_blocks_ip_and_localhost(self):
        assert is_allowed_download_url("http://127.0.0.1/a.pdf") is False
        assert is_allowed_download_url("http://localhost/a.pdf") is False


class TestSniffAndHtmlDetection:
    def test_sniffs_pdf(self):
        assert sniff_extension_from_bytes(b"%PDF-1.7\\n%...") == ".pdf"

    def test_sniffs_zip(self):
        assert sniff_extension_from_bytes(b"PK\x03\x04\x14\x00") == ".zip"

    def test_detects_html(self):
        assert (
            looks_like_html_bytes(b"<!DOCTYPE html><html><head></head></html>") is True
        )
        assert looks_like_html_bytes(b"  <html><title>x</title></html>") is True

    def test_does_not_flag_pdf_as_html(self):
        assert looks_like_html_bytes(b"%PDF-1.7\\n%...") is False


# ---------------------------------------------------------------------------
# pagination_button_is_disabled
# ---------------------------------------------------------------------------


class TestPaginationButtonIsDisabled:
    def test_disabled_property_true(self):
        assert pagination_button_is_disabled(None, True, "") is True

    def test_aria_disabled_true(self):
        assert pagination_button_is_disabled("true", None, "") is True

    def test_class_contains_disabled(self):
        assert pagination_button_is_disabled(None, None, "page-item disabled") is True

    def test_not_disabled(self):
        assert pagination_button_is_disabled(None, False, "page-item") is False


# ---------------------------------------------------------------------------
# status_matches_target (whitespace-insensitive)
# ---------------------------------------------------------------------------


class TestStatusMatchesTarget:
    def test_matches_exact(self):
        assert status_matches_target("หนังสือเชิญชวน/ประกาศเชิญชวน") is True

    def test_matches_with_extra_whitespace(self):
        assert status_matches_target("หนังสือเชิญชวน /\n ประกาศเชิญชวน") is True


class TestShowHtmlfileDetection:
    def test_detects_proc_id_query(self):
        url = (
            "https://process3.gprocurement.go.th/egp2procmainWeb/jsp/procsearch.sch"
            "?pid=1&servlet=gojsp&proc_id=ShowHTMLFile&processFlows=Procure"
        )
        assert is_show_htmlfile_url(url) is True

    def test_non_show_htmlfile_is_false(self):
        assert is_show_htmlfile_url("https://example.com/a.pdf") is False


class TestExtractUrlFromOnclick:
    def test_extracts_absolute_url(self):
        onclick = "window.open('https://process3.gprocurement.go.th/x?proc_id=ShowHTMLFile','_blank')"
        assert extract_url_from_onclick(
            onclick, base_url="https://process5.gprocurement.go.th/"
        ) == ("https://process3.gprocurement.go.th/x?proc_id=ShowHTMLFile")

    def test_extracts_relative_url(self):
        onclick = "window.open('/egp2procmainWeb/jsp/procsearch.sch?proc_id=ShowHTMLFile','_blank')"
        assert extract_url_from_onclick(
            onclick, base_url="https://process3.gprocurement.go.th/egp"
        ) == (
            "https://process3.gprocurement.go.th/egp2procmainWeb/jsp/procsearch.sch?proc_id=ShowHTMLFile"
        )


# ---------------------------------------------------------------------------
# Timeout constants
# ---------------------------------------------------------------------------


class TestTimeoutConstants:
    def test_subpage_timeout_longer_than_regular(self):
        """Sub-page TOR downloads need more time than regular downloads."""
        assert SUBPAGE_DOWNLOAD_TIMEOUT > DOWNLOAD_TIMEOUT

    def test_subpage_timeout_at_least_60s(self):
        assert SUBPAGE_DOWNLOAD_TIMEOUT >= 60_000


class TestSiteErrorToastText:
    def test_matches_full_toast_text(self):
        assert has_site_error_toast_text("ระบบเกิดข้อผิดพลาด กรุณาตรวจสอบ") is True

    def test_matches_multiline_toast_text(self):
        text = "ระบบเกิดข้อผิดพลาด\nกรุณา\nตรวจสอบ"
        assert has_site_error_toast_text(text) is True

    def test_non_error_text_returns_false(self):
        assert has_site_error_toast_text("ดาวน์โหลดสำเร็จ") is False


# ---------------------------------------------------------------------------
# safe_shutdown
# ---------------------------------------------------------------------------


class TestSafeShutdown:
    def test_calls_pw_stop_even_if_browser_close_raises(self):
        calls: list[str] = []

        class DummyBrowser:
            def close(self):
                calls.append("browser.close")
                raise KeyboardInterrupt()

        class DummyPW:
            def stop(self):
                calls.append("pw.stop")

        safe_shutdown(
            browser=DummyBrowser(),
            pw=DummyPW(),
            chrome_proc=None,
            ignore_sigint=False,
        )
        assert calls == ["browser.close", "pw.stop"]

    def test_handles_chrome_proc_terminate_then_kill(self):
        calls: list[str] = []

        class DummyProc:
            def send_signal(self, sig):
                calls.append("send_signal")
                raise RuntimeError("nope")

            def wait(self, timeout: int):
                calls.append("wait")

            def kill(self):
                calls.append("kill")

        safe_shutdown(
            browser=None, pw=None, chrome_proc=DummyProc(), ignore_sigint=False
        )
        assert calls == ["send_signal", "kill", "wait"]


# ---------------------------------------------------------------------------
# Env config helpers
# ---------------------------------------------------------------------------


class TestEnvHelpers:
    def test_env_get_str_defaults_when_missing(self, monkeypatch):
        monkeypatch.delenv("X_STR", raising=False)
        assert env_get_str("X_STR", "default") == "default"

    def test_env_get_str_ignores_empty(self, monkeypatch):
        monkeypatch.setenv("X_STR", "   ")
        assert env_get_str("X_STR", "default") == "default"

    def test_env_get_int_parses_and_validates_range(self, monkeypatch):
        monkeypatch.setenv("X_INT", "10")
        assert env_get_int("X_INT", 5, min_value=1, max_value=20) == 10

        monkeypatch.setenv("X_INT", "0")
        assert env_get_int("X_INT", 5, min_value=1, max_value=20) == 5

    def test_env_get_int_defaults_on_invalid(self, monkeypatch):
        monkeypatch.setenv("X_INT", "nope")
        assert env_get_int("X_INT", 7) == 7

    def test_env_get_float_parses(self, monkeypatch):
        monkeypatch.setenv("X_FLOAT", "1.25")
        assert env_get_float("X_FLOAT", 0.5, min_value=0.0, max_value=2.0) == 1.25

    def test_env_get_path_expands_user(self, monkeypatch):
        monkeypatch.setenv("X_PATH", "~/tmp-egp")
        p = env_get_path("X_PATH", Path("/default"))
        assert str(p).endswith("/tmp-egp")
        assert p.is_absolute()

    def test_parse_keywords_splits_comma_and_newline(self):
        assert parse_keywords("a,b\nc\r\nd,,") == ["a", "b", "c", "d"]

    def test_keywords_from_env_falls_back_when_empty(self, monkeypatch):
        monkeypatch.setenv("X_KW", "   ")
        assert keywords_from_env("X_KW", ["x", "y"]) == ["x", "y"]

    def test_load_dotenv_file_parses_and_sets_env(self, tmp_path, monkeypatch):
        import os

        monkeypatch.delenv("DOTENV_A", raising=False)
        monkeypatch.delenv("DOTENV_B", raising=False)
        monkeypatch.delenv("DOTENV_C", raising=False)

        p = tmp_path / ".env"
        p.write_text(
            "\n".join(
                [
                    "# comment",
                    "DOTENV_A=hello",
                    'DOTENV_B="a b c"',
                    "export DOTENV_C=ok # trailing comment",
                ]
            ),
            encoding="utf-8",
        )

        parsed = load_dotenv_file(p, override=False)
        assert parsed["DOTENV_A"] == "hello"
        assert parsed["DOTENV_B"] == "a b c"
        assert parsed["DOTENV_C"] == "ok"
        assert os.environ["DOTENV_A"] == "hello"
        assert os.environ["DOTENV_B"] == "a b c"
        assert os.environ["DOTENV_C"] == "ok"

    def test_load_dotenv_file_does_not_override_by_default(self, tmp_path, monkeypatch):
        import os

        monkeypatch.setenv("DOTENV_X", "keep")
        p = tmp_path / ".env"
        p.write_text("DOTENV_X=replace\n", encoding="utf-8")
        load_dotenv_file(p, override=False)
        assert os.environ["DOTENV_X"] == "keep"

    def test_load_dotenv_file_override_true_replaces(self, tmp_path, monkeypatch):
        import os

        monkeypatch.setenv("DOTENV_Y", "old")
        p = tmp_path / ".env"
        p.write_text("DOTENV_Y=new\n", encoding="utf-8")
        load_dotenv_file(p, override=True)
        assert os.environ["DOTENV_Y"] == "new"


class TestWaitForLocalTcpListen:
    def test_returns_true_when_listening(self):
        srv = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        srv.bind(("127.0.0.1", 0))
        srv.listen(1)
        port = srv.getsockname()[1]
        try:
            assert (
                wait_for_local_tcp_listen("127.0.0.1", port, timeout_seconds=1.0)
                is True
            )
        finally:
            srv.close()

    def test_returns_false_when_bound_but_not_listening(self):
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.bind(("127.0.0.1", 0))
        port = s.getsockname()[1]
        try:
            assert (
                wait_for_local_tcp_listen("127.0.0.1", port, timeout_seconds=0.5)
                is False
            )
        finally:
            s.close()


class _FakeNavigationPage:
    def __init__(self, *, dom_timeouts: int, commit_timeout: bool = False) -> None:
        self.dom_timeouts = dom_timeouts
        self.commit_timeout = commit_timeout
        self.goto_calls = []

    def goto(self, url: str, wait_until=None, timeout=None):
        self.goto_calls.append((url, wait_until, timeout))
        if wait_until == "domcontentloaded" and self.dom_timeouts > 0:
            self.dom_timeouts -= 1
            raise PlaywrightTimeout("slow DOM")
        if wait_until == "commit" and self.commit_timeout:
            raise PlaywrightTimeout("slow commit")
        return None


class TestGotoWithRetries:
    def test_retries_domcontentloaded_timeout(self, monkeypatch):
        page = _FakeNavigationPage(dom_timeouts=1)
        sleeps = []

        monkeypatch.setattr(
            "egp_crawler.logged_sleep",
            lambda seconds, reason="": sleeps.append((seconds, reason)),
        )

        _goto_with_retries(
            page,
            "https://example.test/start",
            label="startup",
            attempts=2,
            timeout_ms=123,
        )

        assert page.goto_calls == [
            ("https://example.test/start", "domcontentloaded", 123),
            ("https://example.test/start", "domcontentloaded", 123),
        ]
        assert sleeps == [(3, "retry startup navigation")]

    def test_uses_commit_fallback_after_domcontentloaded_retries(self, monkeypatch):
        page = _FakeNavigationPage(dom_timeouts=2)
        sleeps = []

        monkeypatch.setattr(
            "egp_crawler.logged_sleep",
            lambda seconds, reason="": sleeps.append((seconds, reason)),
        )

        _goto_with_retries(
            page,
            "https://example.test/start",
            label="startup",
            attempts=2,
            timeout_ms=60_000,
        )

        assert page.goto_calls == [
            ("https://example.test/start", "domcontentloaded", 60_000),
            ("https://example.test/start", "domcontentloaded", 60_000),
            ("https://example.test/start", "commit", 15_000),
        ]
        assert sleeps == [(3, "retry startup navigation")]

    def test_raises_when_commit_fallback_times_out(self, monkeypatch):
        import pytest

        page = _FakeNavigationPage(dom_timeouts=1, commit_timeout=True)

        monkeypatch.setattr("egp_crawler.logged_sleep", lambda *args, **kwargs: None)

        with pytest.raises(PlaywrightTimeout, match="slow DOM"):
            _goto_with_retries(
                page,
                "https://example.test/start",
                label="startup",
                attempts=1,
                timeout_ms=60_000,
            )


class _FakeTextElement:
    def __init__(self, text: str):
        self._text = text

    def inner_text(self):
        return self._text

    def click(self):
        return None


class _FakeCell:
    def __init__(self, text: str = "", clickable=None):
        self._text = text
        self._clickable = clickable

    def inner_text(self):
        return self._text

    def query_selector(self, selector: str):
        if self._clickable and any(
            key in selector
            for key in ("a", "button", "[role='button']", '[role="button"]', "svg", "i")
        ):
            return self._clickable
        return None

    def click(self):
        if self._clickable:
            self._clickable.click()


class _FakeRow:
    def __init__(self, cells):
        self._cells = cells

    def query_selector_all(self, selector: str):
        if selector == "td":
            return self._cells
        return []

    def inner_text(self):
        return " ".join(cell.inner_text() for cell in self._cells)


class _FakeTable:
    def __init__(self, headers, rows):
        self._headers = headers
        self._rows = rows

    def query_selector_all(self, selector: str):
        if selector in ("thead th, thead td", "th", "tr:first-child th, tr:first-child td"):
            return [_FakeTextElement(h) for h in self._headers]
        if selector == "tbody tr":
            return self._rows
        return []

    def inner_text(self):
        header_text = " ".join(self._headers)
        row_text = " ".join(row.inner_text() for row in self._rows)
        return f"{header_text} {row_text}".strip()


class _FakeResultsPage:
    def __init__(self, tables, body_text: str = "", active_page: str | None = None):
        self._tables = tables
        self._body_text = body_text
        self._active_page = active_page

    def query_selector_all(self, selector: str):
        if selector == "table":
            return self._tables
        if selector == "li.page-item.active, li.active, .pagination .active":
            return [_FakeTextElement(self._active_page)] if self._active_page else []
        return []

    def query_selector(self, selector: str):
        if selector == "li.page-item.active, li.active, .pagination .active":
            return _FakeTextElement(self._active_page) if self._active_page else None
        return None

    def inner_text(self, selector: str):
        if selector == "body":
            return self._body_text
        raise AssertionError(f"unexpected selector: {selector}")


class TestResultsDebugSnapshot:
    def test_includes_results_headers_rows_and_body_snippet(self):
        results_table = _FakeTable(
            ["ลำดับ", "หน่วยจัดซื้อ", "ชื่อโครงการ", "วงเงินงบประมาณ (บาท)", "สถานะโครงการ", "ดูข้อมูล"],
            [
                _FakeRow(
                    [
                        _FakeCell("1"),
                        _FakeCell("หน่วยงาน A"),
                        _FakeCell("โครงการระบบวิเคราะห์"),
                        _FakeCell("100.00"),
                        _FakeCell("หนังสือเชิญชวน/ประกาศเชิญชวน"),
                        _FakeCell("article"),
                    ]
                )
            ],
        )
        page = _FakeResultsPage(
            [results_table],
            body_text="จำนวนโครงการที่พบ : 1 โครงการ\nโครงการระบบวิเคราะห์\nเพิ่มเติม",
        )
        page.url = "https://process5.gprocurement.go.th/egp-agpc01-web/announcement"

        snapshot = build_results_debug_snapshot(page)

        assert snapshot["url"] == page.url
        assert snapshot["table_count"] == 1
        assert snapshot["results_row_count"] == 1
        assert snapshot["results_headers"] == [
            "ลำดับ",
            "หน่วยจัดซื้อ",
            "ชื่อโครงการ",
            "วงเงินงบประมาณ (บาท)",
            "สถานะโครงการ",
            "ดูข้อมูล",
        ]
        assert snapshot["results_row_samples"] == [
            [
                "1",
                "หน่วยงาน A",
                "โครงการระบบวิเคราะห์",
                "100.00",
                "หนังสือเชิญชวน/ประกาศเชิญชวน",
                "article",
            ]
        ]
        assert "จำนวนโครงการที่พบ : 1 โครงการ" in snapshot["body_snippet"]


class TestResultsPageMarker:
    def test_captures_active_page_and_row_sample(self):
        results_table = _FakeTable(
            ["ลำดับ", "หน่วยจัดซื้อ", "ชื่อโครงการ", "วงเงินงบประมาณ (บาท)", "สถานะโครงการ", "ดูข้อมูล"],
            [
                _FakeRow(
                    [
                        _FakeCell("1"),
                        _FakeCell("หน่วยงาน A"),
                        _FakeCell("โครงการ A"),
                        _FakeCell("100.00"),
                        _FakeCell("หนังสือเชิญชวน/ประกาศเชิญชวน"),
                        _FakeCell("article"),
                    ]
                ),
                _FakeRow(
                    [
                        _FakeCell("2"),
                        _FakeCell("หน่วยงาน B"),
                        _FakeCell("โครงการ B"),
                        _FakeCell("200.00"),
                        _FakeCell("จัดทำสัญญา/บริหารสัญญา"),
                        _FakeCell("article"),
                    ]
                ),
            ],
        )
        page = _FakeResultsPage([results_table], active_page="4")

        marker = get_results_page_marker(page)

        assert marker["active_page"] == "4"
        assert marker["row_count"] == 2
        assert "โครงการ A" in marker["row_sample"]
        assert "โครงการ B" in marker["row_sample"]

    def test_detects_page_marker_change(self):
        previous = {"active_page": "1", "row_count": 3, "row_sample": "old"}
        current = {"active_page": "2", "row_count": 3, "row_sample": "old"}

        assert results_page_marker_changed(previous, current) is True


class _FakeWaitPage:
    def __init__(self):
        self.wait_for_selector_calls = []

    def wait_for_selector(self, selector: str, timeout=None, state=None):
        self.wait_for_selector_calls.append(
            {"selector": selector, "timeout": timeout, "state": state}
        )
        return object()

    def wait_for_function(self, expression: str, timeout=None):
        return None

    def query_selector_all(self, selector: str):
        return []

    def query_selector(self, selector: str):
        return None


class TestWaitForResultsReady:
    def test_waits_for_attached_table_not_visible_table(self, monkeypatch):
        page = _FakeWaitPage()
        monkeypatch.setattr("egp_crawler.logged_sleep", lambda *args, **kwargs: None)

        wait_for_results_ready(page)

        assert page.wait_for_selector_calls[0]["selector"] == "table"
        assert page.wait_for_selector_calls[0]["state"] == "attached"


class _FakeCloudflareButton:
    def __init__(self, *, disabled: bool):
        self._disabled = disabled

    def get_attribute(self, name: str):
        if name == "disabled":
            return "" if self._disabled else None
        return None


class _FakeCloudflarePage:
    def __init__(self, *, enabled_after_reload: bool = True):
        self.enabled_after_reload = enabled_after_reload
        self.reload_calls = 0
        self.url = "https://example.test/announcement"

    def query_selector(self, selector: str):
        if "button:has-text('ค้นหา')" in selector:
            disabled = self.reload_calls == 0 or not self.enabled_after_reload
            return _FakeCloudflareButton(disabled=disabled)
        if "iframe[src*='challenges.cloudflare.com']" in selector:
            return object()
        return None

    def reload(self, wait_until=None, timeout=None):
        self.reload_calls += 1

    def goto(self, url: str, wait_until=None, timeout=None):
        self.reload_calls += 1


class TestWaitForCloudflare:
    def test_reloads_once_and_then_passes(self, monkeypatch):
        page = _FakeCloudflarePage(enabled_after_reload=True)
        clock = iter([0.0, 0.0, 1.0, 2.0, 2.0])

        monkeypatch.setattr("egp_crawler.logged_sleep", lambda *args, **kwargs: None)
        monkeypatch.setattr("egp_crawler.time.time", lambda: next(clock))

        wait_for_cloudflare(page, timeout_ms=500, reload_retries=1)

        assert page.reload_calls == 1

    def test_continues_after_reload_budget_exhausted(self, monkeypatch):
        page = _FakeCloudflarePage(enabled_after_reload=False)
        clock = iter([0.0, 0.0, 1.0])

        monkeypatch.setattr("egp_crawler.logged_sleep", lambda *args, **kwargs: None)
        monkeypatch.setattr("egp_crawler.time.time", lambda: next(clock))

        wait_for_cloudflare(page, timeout_ms=500, reload_retries=0)

        assert page.reload_calls == 0


class _FakeSearchInput:
    def __init__(self):
        self.values = []

    def click(self):
        return None

    def fill(self, value: str):
        self.values.append(value)


class _FakeSearchPage:
    def __init__(self):
        self.goto_calls = []
        self.url = "https://process5.gprocurement.go.th/egp-agpc01-web/announcement"
        self._button = _FakeTextElement("ค้นหา")

    def query_selector(self, selector: str):
        if "button:has-text('ค้นหา')" in selector:
            return self._button
        return None

    def wait_for_selector(self, selector: str, timeout=None, state=None):
        if "button:has-text('ค้นหา')" in selector:
            return self._button
        raise AssertionError(f"unexpected selector: {selector}")

    def goto(self, url: str, wait_until=None, timeout=None):
        self.goto_calls.append((url, wait_until, timeout))
        self.url = url


class _FakeNextButton:
    def __init__(self, page) -> None:
        self.page = page

    def is_visible(self) -> bool:
        return self.page.remaining_clicks > 0

    def click(self, timeout=None) -> None:
        self.page.remaining_clicks -= 1


class _FakeRestorePage:
    def __init__(self, pages_to_advance: int) -> None:
        self.remaining_clicks = pages_to_advance
        self.evaluate_calls = 0

    def query_selector(self, selector: str):
        if "ถัดไป" in selector:
            return _FakeNextButton(self)
        return None

    def evaluate(self, script: str, arg=None):
        self.evaluate_calls += 1
        if callable(getattr(arg, "click", None)):
            arg.click()
        return None


class TestSearchKeyword:
    def test_retries_from_fresh_search_page_after_cloudflare_exhausted(
        self, monkeypatch
    ):
        page = _FakeSearchPage()
        search_input = _FakeSearchInput()
        cloudflare_results = iter([False, True])
        waits = []

        monkeypatch.setattr(
            "egp_crawler.wait_for_cloudflare",
            lambda page, timeout_ms=None, reload_retries=None: next(cloudflare_results),
        )
        monkeypatch.setattr("egp_crawler.find_search_input", lambda page, btn: search_input)
        monkeypatch.setattr("egp_crawler.click_search_button", lambda page, btn=None: None)
        monkeypatch.setattr("egp_crawler.wait_for_results_ready", lambda page: None)
        monkeypatch.setattr("egp_crawler.get_results_rows", lambda page: [])
        monkeypatch.setattr(
            "egp_crawler.logged_sleep",
            lambda seconds, reason="": waits.append((seconds, reason)),
        )

        search_keyword(page, "ธรรมาภิบาลข้อมูล")

        assert page.goto_calls == [(SEARCH_URL, "domcontentloaded", 60_000)]
        assert search_input.values == ["", "ธรรมาภิบาลข้อมูล"]

    def test_waits_for_full_row_stabilization_before_reporting_count(
        self, monkeypatch
    ):
        page = _FakeSearchPage()
        search_input = _FakeSearchInput()
        waits = []
        row_counts = iter([1, 1, 10, 10, 10])
        printed = []

        monkeypatch.setattr("egp_crawler.wait_for_cloudflare", lambda page: None)
        monkeypatch.setattr("egp_crawler.find_search_input", lambda page, btn: search_input)
        monkeypatch.setattr("egp_crawler.wait_for_results_ready", lambda page: None)
        monkeypatch.setattr(
            "egp_crawler.get_results_rows",
            lambda page: [object()] * next(row_counts, 10),
        )
        monkeypatch.setattr(
            "egp_crawler.logged_sleep",
            lambda seconds, reason="": waits.append((seconds, reason)),
        )
        monkeypatch.setattr("egp_crawler.print", lambda *args, **kwargs: printed.append(args))

        search_keyword(page, "ระบบวิเคราะห์")

        assert search_input.values == ["", "ระบบวิเคราะห์"]
        assert any("rows: 10" in " ".join(str(part) for part in line) for line in printed)

    def test_restore_results_page_replays_search_and_advances_pages(self, monkeypatch):
        page = _FakeRestorePage(pages_to_advance=2)
        search_calls = []
        marker_values = iter(
            [
                {"active_page": "1", "row_count": 10, "row_sample": "a"},
                {"active_page": "2", "row_count": 10, "row_sample": "b"},
            ]
        )

        monkeypatch.setattr(
            "egp_crawler.search_keyword",
            lambda page, keyword: search_calls.append(keyword),
        )
        monkeypatch.setattr("egp_crawler.dismiss_modal", lambda page: None)
        monkeypatch.setattr(
            "egp_crawler.get_results_page_marker", lambda page: next(marker_values)
        )
        monkeypatch.setattr(
            "egp_crawler.wait_for_results_page_change", lambda page, previous_marker: True
        )
        monkeypatch.setattr("egp_crawler.logged_sleep", lambda *args, **kwargs: None)

        restore_results_page(page, "ระบบวิเคราะห์", 3)

        assert search_calls == ["ระบบวิเคราะห์"]
        assert page.remaining_clicks == 0


# ---------------------------------------------------------------------------
# parse_cli_args
# ---------------------------------------------------------------------------


class TestParseCliArgs:
    def test_default_profile_is_tor(self):
        args = parse_cli_args([])
        assert args.profile == "tor"

    def test_explicit_tor_profile(self):
        args = parse_cli_args(["--profile", "tor"])
        assert args.profile == "tor"

    def test_explicit_toe_profile(self):
        args = parse_cli_args(["--profile", "toe"])
        assert args.profile == "toe"

    def test_explicit_lue_profile(self):
        args = parse_cli_args(["--profile", "lue"])
        assert args.profile == "lue"

    def test_invalid_profile_raises(self):
        import pytest
        with pytest.raises(SystemExit):
            parse_cli_args(["--profile", "invalid"])


# ---------------------------------------------------------------------------
# apply_profile_defaults
# ---------------------------------------------------------------------------


class TestApplyProfileDefaults:
    def _clear_profile_env(self):
        for key in ("EGP_KEYWORDS", "EGP_DOWNLOAD_DIR", "EGP_LOCAL_FALLBACK_DIR"):
            os.environ.pop(key, None)

    def test_tor_profile_sets_tor_keywords(self):
        self._clear_profile_env()
        apply_profile_defaults("tor")
        kws = [k.strip() for k in os.environ["EGP_KEYWORDS"].split(",")]
        assert kws == KEYWORDS_DEFAULT
        self._clear_profile_env()

    def test_toe_profile_sets_toe_keywords(self):
        self._clear_profile_env()
        apply_profile_defaults("toe")
        kws = [k.strip() for k in os.environ["EGP_KEYWORDS"].split(",")]
        assert kws == KEYWORDS_TOE_DEFAULT
        self._clear_profile_env()

    def test_tor_profile_sets_tor_download_dir(self):
        self._clear_profile_env()
        apply_profile_defaults("tor")
        assert os.environ["EGP_DOWNLOAD_DIR"].endswith("TOR")
        self._clear_profile_env()

    def test_toe_profile_sets_toe_download_dir(self):
        self._clear_profile_env()
        apply_profile_defaults("toe")
        assert os.environ["EGP_DOWNLOAD_DIR"].endswith("TOE")
        self._clear_profile_env()

    def test_lue_profile_sets_lue_keywords(self):
        self._clear_profile_env()
        apply_profile_defaults("lue")
        kws = [k.strip() for k in os.environ["EGP_KEYWORDS"].split(",")]
        assert kws == KEYWORDS_LUE_DEFAULT
        self._clear_profile_env()

    def test_lue_profile_sets_lue_download_dir(self):
        self._clear_profile_env()
        apply_profile_defaults("lue")
        assert os.environ["EGP_DOWNLOAD_DIR"].endswith("LUE")
        self._clear_profile_env()

    def test_does_not_override_existing_shell_env(self):
        self._clear_profile_env()
        os.environ["EGP_DOWNLOAD_DIR"] = "/custom/path"
        apply_profile_defaults("toe")
        assert os.environ["EGP_DOWNLOAD_DIR"] == "/custom/path"
        self._clear_profile_env()


# ---------------------------------------------------------------------------
# PROFILE_DEFAULTS
# ---------------------------------------------------------------------------


class TestProfileDefaults:
    def test_tor_and_toe_profiles_exist(self):
        assert "tor" in PROFILE_DEFAULTS
        assert "toe" in PROFILE_DEFAULTS

    def test_lue_profile_exists(self):
        assert "lue" in PROFILE_DEFAULTS

    def test_tor_keywords_match_keywords_default(self):
        assert PROFILE_DEFAULTS["tor"]["keywords"] == KEYWORDS_DEFAULT

    def test_toe_keywords_match_toe_default(self):
        assert PROFILE_DEFAULTS["toe"]["keywords"] == KEYWORDS_TOE_DEFAULT

    def test_lue_keywords_match_lue_default(self):
        assert PROFILE_DEFAULTS["lue"]["keywords"] == KEYWORDS_LUE_DEFAULT

    def test_toe_has_five_keywords(self):
        assert len(KEYWORDS_TOE_DEFAULT) == 5

    def test_lue_has_five_keywords(self):
        assert len(KEYWORDS_LUE_DEFAULT) == 5

    def test_profiles_use_different_dirs(self):
        assert PROFILE_DEFAULTS["tor"]["download_dir"] != PROFILE_DEFAULTS["toe"]["download_dir"]

    def test_lue_dir_differs_from_tor_and_toe(self):
        assert PROFILE_DEFAULTS["lue"]["download_dir"] != PROFILE_DEFAULTS["tor"]["download_dir"]
        assert PROFILE_DEFAULTS["lue"]["download_dir"] != PROFILE_DEFAULTS["toe"]["download_dir"]

    def test_toe_download_dir_ends_with_toe(self):
        assert str(PROFILE_DEFAULTS["toe"]["download_dir"]).endswith("TOE")

    def test_tor_download_dir_ends_with_tor(self):
        assert str(PROFILE_DEFAULTS["tor"]["download_dir"]).endswith("TOR")

    def test_lue_download_dir_ends_with_lue(self):
        assert str(PROFILE_DEFAULTS["lue"]["download_dir"]).endswith("LUE")


class TestProjectDocumentDownloads:
    def test_doc_targets_include_final_tor(self):
        assert "เอกสารประกวดราคา" in DOCS_TO_DOWNLOAD

    def test_returns_draft_bucket_when_only_draft_tor_is_downloaded(
        self, monkeypatch, tmp_path
    ):
        monkeypatch.setattr(
            "egp_crawler._download_one_document",
            lambda page, target_doc, project_dir: {
                "ประกาศเชิญชวน": [],
                "ประกาศราคากลาง": ["ประกาศราคากลาง"],
                "ร่างเอกสารประกวดราคา": ["ร่างเอกสารประกวดราคา"],
                "เอกสารประกวดราคา": [],
            }[target_doc],
        )

        summary = download_project_documents(_FakeResultsPage([]), tmp_path)

        assert summary.tor_downloaded is False
        assert summary.artifact_bucket is ArtifactBucket.DRAFT_PLUS_PRICING

    def test_invitation_popup_counts_final_tor_download(self, monkeypatch, tmp_path):
        clickable = object()
        page = _FakeResultsPage(
            [
                _FakeTable(
                    ["ลำดับ", "ประกาศที่เกี่ยวข้อง", "วันที่ประกาศ", "ดูข้อมูล"],
                    [
                        _FakeRow(
                            [
                                _FakeCell("1"),
                                _FakeCell("ประกาศเชิญชวน"),
                                _FakeCell("10/04/2569"),
                                _FakeCell("", clickable=clickable),
                            ]
                        )
                    ],
                )
            ]
        )

        monkeypatch.setattr("egp_crawler.dismiss_modal", lambda page: None)
        monkeypatch.setattr("egp_crawler.logged_sleep", lambda *args, **kwargs: None)
        monkeypatch.setattr(
            "egp_crawler._handle_direct_or_page_download",
            lambda page, btn, project_dir, doc_name: None,
        )
        monkeypatch.setattr(
            "egp_crawler._download_documents_from_current_view",
            lambda page, project_dir, include_label: [
                "ประกาศเชิญชวน",
                "เอกสารประกวดราคา",
            ],
            raising=False,
        )

        assert _download_one_document(page, "ประกาศเชิญชวน", tmp_path) == [
            "ประกาศเชิญชวน",
            "เอกสารประกวดราคา",
        ]

    def test_write_project_manifest_records_saved_files(self, tmp_path):
        project_dir = tmp_path / "example-project"
        project_dir.mkdir()
        (project_dir / "ประกาศราคากลาง.zip").write_bytes(b"zip")
        manifest_path = write_project_manifest(
            project_dir=project_dir,
            project_info={
                "project_number": "6901",
                "project_name": "Example Project",
                "search_name": "Example Project",
                "keyword": "ระบบสารสนเทศ",
            },
            tracking_status=ProjectState.OPEN_INVITATION,
            closed_reason=None,
            artifact_bucket=ArtifactBucket.PRICING_ONLY,
        )

        payload = json.loads(manifest_path.read_text(encoding="utf-8"))
        assert payload["saved_by"] == "crawler"
        assert payload["artifact_bucket"] == ArtifactBucket.PRICING_ONLY.value
        assert payload["saved_files"] == ["ประกาศราคากลาง.zip"]


class _FakeResponse:
    def __init__(self, body: bytes, *, headers: dict[str, str] | None = None, ok: bool = True):
        self._body = body
        self.headers = headers or {}
        self.ok = ok
        self.status = 200 if ok else 500

    def body(self):
        return self._body


class _FakeRequestClient:
    def __init__(self, response: _FakeResponse):
        self._response = response
        self.calls = []

    def get(self, url: str, timeout=None):
        self.calls.append({"url": url, "timeout": timeout})
        return self._response


class _FakeKeyboard:
    def __init__(self):
        self.keys = []

    def press(self, key: str):
        self.keys.append(key)


class _FakeViewerPage:
    def __init__(self, *, url: str, embedded_src: str, response: _FakeResponse):
        self.url = url
        self._embedded_src = embedded_src
        self.request = _FakeRequestClient(response)
        self.keyboard = _FakeKeyboard()

    def wait_for_load_state(self, state: str, timeout=None):
        return None

    def evaluate(self, script: str):
        if "embed[src]" in script or "iframe[src]" in script or "object[data]" in script:
            return self._embedded_src
        return None


class TestNewTabFallback:
    def test_save_from_new_tab_uses_request_for_blob_viewer(self, monkeypatch, tmp_path):
        viewer_page = _FakeViewerPage(
            url="blob:https://process5.gprocurement.go.th/example-blob",
            embedded_src="https://process5.gprocurement.go.th/egp-download/final-tor.zip",
            response=_FakeResponse(
                b"PK\x03\x04zip-bytes",
                headers={"content-type": "application/zip"},
            ),
        )

        monkeypatch.setattr("egp_crawler.logged_sleep", lambda *args, **kwargs: None)
        monkeypatch.setattr("egp_crawler.run_with_toast_recovery", lambda *args, **kwargs: (_ for _ in ()).throw(Exception("no download event")))
        monkeypatch.setattr("egp_crawler._cancel_pending_downloads", lambda page: None)

        saved_name = _save_from_new_tab(viewer_page, tmp_path, "ประกาศเชิญชวน")

        assert saved_name == "ประกาศเชิญชวน.zip"
        assert (tmp_path / saved_name).read_bytes() == b"PK\x03\x04zip-bytes"
        assert viewer_page.request.calls == [
            {
                "url": "https://process5.gprocurement.go.th/egp-download/final-tor.zip",
                "timeout": SUBPAGE_DOWNLOAD_TIMEOUT,
            }
        ]
