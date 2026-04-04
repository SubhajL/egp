"""Tests for Excel export service."""

from __future__ import annotations

import base64
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from egp_api.main import create_app
from egp_api.services.export_service import ExportService
from egp_db.repositories.project_repo import SqlProjectRepository, build_project_upsert_record
from egp_shared_types.enums import ClosedReason, ProcurementType, ProjectState

TENANT_ID = "11111111-1111-1111-1111-111111111111"


def test_export_to_excel_produces_valid_xlsx_with_headers(tmp_path) -> None:
    database_url = f"sqlite+pysqlite:///{tmp_path / 'export.sqlite3'}"
    repo = SqlProjectRepository(database_url=database_url, bootstrap_schema=True)

    repo.upsert_project(
        build_project_upsert_record(
            tenant_id=TENANT_ID,
            project_number="EGP-2026-0001",
            search_name="ระบบสารสนเทศ",
            detail_name="จัดซื้อระบบสารสนเทศ",
            project_name="จัดซื้อระบบสารสนเทศ",
            organization_name="กรมตัวอย่าง",
            proposal_submission_date="2026-05-01",
            budget_amount="1500000.00",
            procurement_type=ProcurementType.SERVICES,
            project_state=ProjectState.OPEN_INVITATION,
        ),
        source_status_text="ประกาศเชิญชวน",
    )

    service = ExportService(repo)
    excel_bytes = service.export_to_excel(tenant_id=TENANT_ID)

    assert len(excel_bytes) > 0
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    assert ws is not None

    # Check header row has all 10 columns
    headers = [ws.cell(row=1, column=i).value for i in range(1, 11)]
    assert headers[0] == "ชื่อโครงการ"
    assert headers[5] == "งบประมาณ (บาท)"
    assert headers[9] == "เหตุผลการปิด"

    # Check data row
    assert ws.cell(row=2, column=1).value == "จัดซื้อระบบสารสนเทศ"
    assert ws.cell(row=2, column=2).value == "กรมตัวอย่าง"
    assert ws.cell(row=2, column=3).value == "EGP-2026-0001"
    assert ws.cell(row=2, column=4).value == "บริการ"
    assert ws.cell(row=2, column=5).value == "เปิดรับข้อเสนอ"
    budget_value = ws.cell(row=2, column=6).value
    assert budget_value is not None and float(budget_value) > 0


def test_export_respects_state_filter(tmp_path) -> None:
    database_url = f"sqlite+pysqlite:///{tmp_path / 'export.sqlite3'}"
    repo = SqlProjectRepository(database_url=database_url, bootstrap_schema=True)

    for i, state in enumerate([ProjectState.OPEN_INVITATION, ProjectState.DISCOVERED]):
        repo.upsert_project(
            build_project_upsert_record(
                tenant_id=TENANT_ID,
                project_number=f"EGP-2026-{i:04d}",
                search_name=f"โครงการ {i}",
                detail_name=f"โครงการ {i}",
                project_name=f"โครงการ {i}",
                organization_name="กรม",
                proposal_submission_date="2026-05-01",
                budget_amount="1000000",
                procurement_type=ProcurementType.SERVICES,
                project_state=state,
            ),
        )

    service = ExportService(repo)
    excel_bytes = service.export_to_excel(
        tenant_id=TENANT_ID,
        project_states=["discovered"],
    )
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    assert ws is not None
    assert ws.max_row == 2  # header + 1 filtered row
    assert ws.cell(row=2, column=5).value == "ค้นพบใหม่"


def _seed_project(
    client: TestClient,
    *,
    project_number: str,
    project_name: str,
    organization_name: str,
    procurement_type: ProcurementType,
    project_state: ProjectState,
    budget_amount: str,
    closed_reason: ClosedReason | None = None,
):
    repository = client.app.state.project_repository
    return repository.upsert_project(
        build_project_upsert_record(
            tenant_id=TENANT_ID,
            project_number=project_number,
            search_name=project_name,
            detail_name=project_name,
            project_name=project_name,
            organization_name=organization_name,
            proposal_submission_date="2026-05-01",
            budget_amount=budget_amount,
            procurement_type=procurement_type,
            project_state=project_state,
            closed_reason=closed_reason,
        ),
        source_status_text=project_name,
    )


def _ingest_document(
    client: TestClient,
    *,
    project_id: str,
    file_name: str,
    content: bytes,
    source_label: str,
    source_status_text: str,
) -> None:
    response = client.post(
        "/v1/documents/ingest",
        json={
            "tenant_id": TENANT_ID,
            "project_id": project_id,
            "file_name": file_name,
            "content_base64": base64.b64encode(content).decode("ascii"),
            "source_label": source_label,
            "source_status_text": source_status_text,
        },
    )
    assert response.status_code == 201


def _exported_project_numbers(excel_bytes: bytes) -> list[str]:
    workbook = load_workbook(BytesIO(excel_bytes))
    worksheet = workbook.active
    assert worksheet is not None
    values: list[str] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        project_number = row[2]
        if isinstance(project_number, str) and project_number:
            values.append(project_number)
    return values


def test_export_route_matches_explorer_filter_contract(tmp_path) -> None:
    database_url = f"sqlite+pysqlite:///{tmp_path / 'export-route.sqlite3'}"
    client = TestClient(create_app(artifact_root=tmp_path, database_url=database_url, auth_required=False))

    matched = _seed_project(
        client,
        project_number="EGP-2026-3101",
        project_name="ที่ปรึกษาระบบสุขภาพดิจิทัล",
        organization_name="กระทรวงสาธารณสุข",
        procurement_type=ProcurementType.CONSULTING,
        project_state=ProjectState.WINNER_ANNOUNCED,
        budget_amount="3200000",
        closed_reason=ClosedReason.WINNER_ANNOUNCED,
    )
    _seed_project(
        client,
        project_number="EGP-2026-3102",
        project_name="จัดซื้อเครื่องแม่ข่ายส่วนกลาง",
        organization_name="เทศบาลตัวอย่าง",
        procurement_type=ProcurementType.GOODS,
        project_state=ProjectState.OPEN_INVITATION,
        budget_amount="800000",
    )
    _seed_project(
        client,
        project_number="EGP-2026-3103",
        project_name="ที่ปรึกษาระบบสุขภาพดิจิทัล ระยะที่ 2",
        organization_name="กระทรวงสาธารณสุข",
        procurement_type=ProcurementType.CONSULTING,
        project_state=ProjectState.DISCOVERED,
        budget_amount="1500000",
    )

    _ingest_document(
        client,
        project_id=matched.id,
        file_name="tor-final-v1.pdf",
        content=b"tor-v1",
        source_label="เอกสารประกวดราคา",
        source_status_text="ประกาศเชิญชวน",
    )
    _ingest_document(
        client,
        project_id=matched.id,
        file_name="tor-final-v2.pdf",
        content=b"tor-v2",
        source_label="เอกสารประกวดราคา",
        source_status_text="ประกาศเชิญชวน",
    )

    params = [
        ("tenant_id", TENANT_ID),
        ("project_state", ProjectState.WINNER_ANNOUNCED.value),
        ("project_state", ProjectState.DISCOVERED.value),
        ("procurement_type", ProcurementType.CONSULTING.value),
        ("keyword", "สุขภาพ"),
        ("budget_min", "2000000"),
        ("budget_max", "4000000"),
        ("has_changed_tor", "true"),
        ("has_winner", "true"),
    ]

    list_response = client.get("/v1/projects", params=params)
    export_response = client.get("/v1/exports/excel", params=params)

    assert list_response.status_code == 200
    assert export_response.status_code == 200
    assert export_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    expected_numbers = [project["project_number"] for project in list_response.json()["projects"]]
    assert expected_numbers == [matched.project_number]
    assert _exported_project_numbers(export_response.content) == expected_numbers


def test_export_route_rejects_invalid_filters(tmp_path) -> None:
    database_url = f"sqlite+pysqlite:///{tmp_path / 'export-route-invalid.sqlite3'}"
    client = TestClient(create_app(artifact_root=tmp_path, database_url=database_url, auth_required=False))

    response = client.get(
        "/v1/exports/excel",
        params={"tenant_id": TENANT_ID, "budget_min": "not-a-number"},
    )

    assert response.status_code == 422
